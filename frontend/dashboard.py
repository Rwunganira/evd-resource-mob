"""
WHO EVD Partner Resource Mobilization Dashboard
Run: streamlit run dashboard.py
"""
import io
import requests
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from datetime import date, datetime

import os
API_BASE = os.getenv("FLASK_URL", "http://localhost:5000").rstrip("/") + "/api"

WHO_BLUE = "#0093D5"
WHO_DARK = "#003D6B"
ALERT_RED = "#CC0000"
WARNING_ORANGE = "#FF6600"
SUCCESS_GREEN = "#008000"

st.set_page_config(
    page_title="WHO EVD Resource Mobilization",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Shared CSS ────────────────────────────────────────────────────────────────
st.markdown(f"""
<style>
  [data-testid="stSidebar"] {{ background-color: {WHO_DARK}; }}
  [data-testid="stSidebar"] * {{ color: white !important; }}
  .metric-card {{
      background: white; border-left: 5px solid {WHO_BLUE};
      border-radius: 6px; padding: 12px 16px;
      box-shadow: 0 2px 6px rgba(0,0,0,.10);
  }}
  .phase-banner {{
      padding: 10px 20px; border-radius: 6px;
      font-size: 1.1rem; font-weight: 700;
      text-align: center; margin-bottom: 12px;
  }}
  .alert-box {{
      background: #fff3cd; border-left: 4px solid #ffc107;
      padding: 8px 14px; border-radius: 4px; margin: 4px 0;
  }}
  .status-deployed  {{ color: {SUCCESS_GREEN}; font-weight: 600; }}
  .status-committed {{ color: #b8860b; font-weight: 600; }}
  .status-pipeline  {{ color: {WARNING_ORANGE}; font-weight: 600; }}
  .status-review    {{ color: {ALERT_RED}; font-weight: 600; }}
</style>
""", unsafe_allow_html=True)


# ── API helpers ───────────────────────────────────────────────────────────────
@st.cache_data(ttl=30)
def api_get(endpoint: str, params: dict | None = None):
    try:
        r = requests.get(f"{API_BASE}/{endpoint.lstrip('/')}", params=params, timeout=5)
        r.raise_for_status()
        return r.json().get("data"), None
    except requests.exceptions.ConnectionError:
        return None, "API unavailable — Flask server is not running."
    except Exception as exc:
        return None, str(exc)


def api_post(endpoint: str, payload: dict):
    try:
        r = requests.post(f"{API_BASE}/{endpoint.lstrip('/')}", json=payload, timeout=5)
        return r.json(), r.status_code
    except requests.exceptions.ConnectionError:
        return {"status": "error", "message": "API unavailable"}, 503
    except Exception as exc:
        return {"status": "error", "message": str(exc)}, 500


def api_put(endpoint: str, payload: dict):
    try:
        r = requests.put(f"{API_BASE}/{endpoint.lstrip('/')}", json=payload, timeout=5)
        return r.json(), r.status_code
    except requests.exceptions.ConnectionError:
        return {"status": "error", "message": "API unavailable"}, 503
    except Exception as exc:
        return {"status": "error", "message": str(exc)}, 500


def api_delete(endpoint: str):
    try:
        r = requests.delete(f"{API_BASE}/{endpoint.lstrip('/')}", timeout=5)
        return r.json(), r.status_code
    except requests.exceptions.ConnectionError:
        return {"status": "error", "message": "API unavailable"}, 503
    except Exception as exc:
        return {"status": "error", "message": str(exc)}, 500


def api_offline_msg(msg: str):
    st.error(f"API Error: {msg}")
    st.info("Make sure Flask is running: `python app.py`")


def fmt_usd(val):
    if val >= 1_000_000:
        return f"${val/1_000_000:.1f}M"
    if val >= 1_000:
        return f"${val/1_000:.0f}K"
    return f"${val:,.0f}"


# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown(f"""
    <div style='text-align:center;padding:10px 0 4px;'>
      <div style='font-size:2rem;'>🏥</div>
      <div style='font-size:1.1rem;font-weight:800;color:white;'>WHO EOC</div>
      <div style='font-size:.75rem;color:#a0c4e8;'>EVD Response — DRC</div>
    </div>
    <hr style='border-color:#1a5276;margin:8px 0;'>
    """, unsafe_allow_html=True)

    page = st.radio(
        "Navigation",
        [
            "Overview / Command Center",
            "Partner Tracker",
            "Resource Mobilization",
            "3W Matrix",
            "Situation Reports",
            "Reports & Export",
            "📋 Government Activities",
            "🤝 Partner Contributions",
        ],
        label_visibility="collapsed",
    )
    st.markdown("<hr style='border-color:#1a5276;'>", unsafe_allow_html=True)
    st.markdown(f"<div style='font-size:.7rem;color:#a0c4e8;'>Last refresh: {datetime.now():%H:%M:%S}</div>", unsafe_allow_html=True)
    if st.button("Refresh Data"):
        st.cache_data.clear()
        st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# PAGE 1 — OVERVIEW / COMMAND CENTER
# ══════════════════════════════════════════════════════════════════════════════
if page == "Overview / Command Center":
    st.title("EVD Response — Command Center")

    summary, err = api_get("dashboard/summary")
    if err:
        api_offline_msg(err)
        st.stop()

    # Outbreak phase banner
    phase = summary.get("current_phase")
    if phase:
        phase_name = phase["phase_name"]
        phase_colors = {
            "Alert": ALERT_RED, "Mobilization": WARNING_ORANGE,
            "Response": "#1a7a1a", "Scale-down": WHO_BLUE, "End": "#555",
        }
        bg = phase_colors.get(phase_name, WHO_BLUE)
        st.markdown(
            f"<div class='phase-banner' style='background:{bg};color:white;'>"
            f"OUTBREAK PHASE: {phase_name.upper()}&nbsp;&nbsp;|&nbsp;&nbsp;"
            f"Day {summary['outbreak_day']} of EVD Response"
            "</div>",
            unsafe_allow_html=True,
        )

    # Key metrics
    col1, col2, col3, col4, col5 = st.columns(5)
    f = summary.get("funding", {})
    p = summary.get("partners", {})
    sr = summary.get("latest_sitrep") or {}

    col1.metric("Partners Mobilized", p.get("total", 0), delta=f"+{p.get('active', 0)} active")
    col2.metric("Funding Committed", fmt_usd(f.get("total_committed_usd", 0)))
    col3.metric("Funding Gap", fmt_usd(f.get("gap_usd", 0)),
                delta=f"{f.get('coverage_percent', 0):.0f}% covered", delta_color="off")
    col4.metric("ETUs Operational", summary.get("etus_operational", 0))
    col5.metric("Outbreak Day", summary.get("outbreak_day", 0))

    st.divider()

    col_left, col_right = st.columns([1.4, 1])

    with col_left:
        st.subheader("Latest Situation Report")
        if sr:
            c1, c2, c3 = st.columns(3)
            c1.metric("Confirmed Cases", sr.get("confirmed_cases", 0))
            c2.metric("Deaths", sr.get("deaths", 0))
            c3.metric("CFR %", f"{sr.get('cfr_percent', 0)}%")
            st.caption(f"Report date: {sr.get('report_date')}  |  HCW affected: {sr.get('healthcare_workers_affected', 0)}")
            if sr.get("notes"):
                st.info(sr["notes"])
        else:
            st.warning("No situation reports available.")

        # Funding coverage gauge
        st.subheader("Funding Coverage")
        coverage = f.get("coverage_percent", 0)
        fig_gauge = go.Figure(go.Indicator(
            mode="gauge+number+delta",
            value=coverage,
            number={"suffix": "%"},
            delta={"reference": 100, "increasing": {"color": SUCCESS_GREEN}},
            gauge={
                "axis": {"range": [0, 100]},
                "bar": {"color": WHO_BLUE},
                "steps": [
                    {"range": [0, 40], "color": "#ffd0d0"},
                    {"range": [40, 75], "color": "#fff0c0"},
                    {"range": [75, 100], "color": "#d0f0d0"},
                ],
                "threshold": {"line": {"color": ALERT_RED, "width": 3}, "value": 80},
            },
            title={"text": "Funding Coverage"},
        ))
        fig_gauge.update_layout(height=220, margin=dict(t=40, b=0, l=20, r=20))
        st.plotly_chart(fig_gauge, use_container_width=True)

    with col_right:
        st.subheader("Alerts")
        due_data, _ = api_get("resources/reporting-due", {"days": 3})
        if due_data:
            overdue = due_data.get("overdue", [])
            due_soon = due_data.get("due_soon", [])
            if overdue:
                st.markdown(f"**{len(overdue)} overdue reports:**")
                for r in overdue:
                    st.markdown(
                        f"<div class='alert-box'>⚠️ <b>{r['partner_name']}</b> — "
                        f"{r['resource_type']} report overdue since {r['next_report_due']}</div>",
                        unsafe_allow_html=True,
                    )
            if due_soon:
                st.markdown(f"**{len(due_soon)} due within 3 days:**")
                for r in due_soon:
                    st.markdown(
                        f"<div class='alert-box' style='background:#e8f4fd;border-color:{WHO_BLUE};'>"
                        f"📋 <b>{r['partner_name']}</b> — due {r['next_report_due']}</div>",
                        unsafe_allow_html=True,
                    )
            if not overdue and not due_soon:
                st.success("No overdue reports in the next 3 days.")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE 2 — PARTNER TRACKER
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Partner Tracker":
    st.title("Partner Tracker")

    partners, err = api_get("partners")
    if err:
        api_offline_msg(err)
        st.stop()

    df = pd.DataFrame(partners)

    if df.empty:
        st.info("No partners found.")
        st.stop()

    # Filters
    col1, col2, col3 = st.columns(3)
    with col1:
        types = ["All"] + sorted(df["partner_type"].unique().tolist())
        sel_type = st.selectbox("Filter by Type", types)
    with col2:
        statuses = ["All"] + sorted(df["status"].unique().tolist())
        sel_status = st.selectbox("Filter by Status", statuses)
    with col3:
        countries = ["All"] + sorted(df["country"].unique().tolist())
        sel_country = st.selectbox("Filter by Country", countries)

    filtered = df.copy()
    if sel_type != "All":
        filtered = filtered[filtered["partner_type"] == sel_type]
    if sel_status != "All":
        filtered = filtered[filtered["status"] == sel_status]
    if sel_country != "All":
        filtered = filtered[filtered["country"] == sel_country]

    st.markdown(f"**{len(filtered)} partners** matching filters")
    st.dataframe(
        filtered[["name", "partner_type", "country", "status", "contact_person", "resource_count", "activity_count"]],
        use_container_width=True, hide_index=True,
        column_config={
            "name": "Partner",
            "partner_type": "Type",
            "country": "Country",
            "status": "Status",
            "contact_person": "Contact",
            "resource_count": "Resources",
            "activity_count": "Activities",
        },
    )

    col_a, col_b = st.columns(2)
    with col_a:
        st.subheader("By Status")
        fig_pie = px.pie(
            df, names="status", title="Partner Status Distribution",
            color_discrete_map={
                "Active": SUCCESS_GREEN, "Negotiating": WHO_BLUE,
                "Pipeline": WARNING_ORANGE, "Inactive": "#aaa",
            },
            hole=0.4,
        )
        fig_pie.update_layout(height=300, margin=dict(t=40, b=0))
        st.plotly_chart(fig_pie, use_container_width=True)

    with col_b:
        st.subheader("By Type")
        type_counts = df["partner_type"].value_counts().reset_index()
        type_counts.columns = ["Type", "Count"]
        fig_bar = px.bar(type_counts, x="Count", y="Type", orientation="h",
                         color_discrete_sequence=[WHO_BLUE])
        fig_bar.update_layout(height=300, margin=dict(t=10, b=0))
        st.plotly_chart(fig_bar, use_container_width=True)

    # Add new partner
    st.divider()
    st.subheader("Add New Partner")
    with st.form("add_partner_form"):
        fc1, fc2 = st.columns(2)
        p_name = fc1.text_input("Partner Name *")
        p_type = fc2.selectbox("Type *", ["Government", "NGO", "Private", "Academic", "Military", "Community"])
        fd1, fd2 = st.columns(2)
        p_country = fd1.text_input("Country *")
        p_status = fd2.selectbox("Status", ["Pipeline", "Negotiating", "Active"])
        fe1, fe2 = st.columns(2)
        p_contact = fe1.text_input("Contact Person")
        p_email = fe2.text_input("Contact Email")
        submitted = st.form_submit_button("Add Partner", type="primary")
        if submitted:
            if not p_name or not p_country:
                st.error("Name and Country are required.")
            else:
                resp, code = api_post("partners", {
                    "name": p_name, "partner_type": p_type, "country": p_country,
                    "status": p_status, "contact_person": p_contact, "contact_email": p_email,
                })
                if code in (200, 201):
                    st.success(f"Partner '{p_name}' added successfully!")
                    st.cache_data.clear()
                    st.rerun()
                else:
                    st.error(resp.get("message", "Error adding partner"))

    # Partner detail
    st.divider()
    st.subheader("Partner Detail")
    partner_names = {p["name"]: p["id"] for p in partners}
    selected_name = st.selectbox("Select partner to view detail", ["— select —"] + list(partner_names))
    if selected_name != "— select —":
        detail, _ = api_get(f"partners/{partner_names[selected_name]}")
        if detail:
            c1, c2 = st.columns(2)
            with c1:
                st.markdown(f"**Type:** {detail['partner_type']}  |  **Country:** {detail['country']}")
                st.markdown(f"**Status:** {detail['status']}  |  **Contact:** {detail.get('contact_person','')} ({detail.get('contact_email','')})")
                st.markdown("**Resources:**")
                if detail["resources"]:
                    st.dataframe(pd.DataFrame(detail["resources"])[["resource_type", "amount", "currency", "status"]],
                                 hide_index=True, use_container_width=True)
            with c2:
                st.markdown("**Activities:**")
                if detail["activities"]:
                    st.dataframe(pd.DataFrame(detail["activities"])[["activity_type", "location", "status"]],
                                 hide_index=True, use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE 3 — RESOURCE MOBILIZATION
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Resource Mobilization":
    st.title("Resource Mobilization")

    resources, err = api_get("resources")
    if err:
        api_offline_msg(err)
        st.stop()

    df = pd.DataFrame(resources)

    gap_data, _ = api_get("resources/funding-gap")
    by_type_data, _ = api_get("resources/total-by-type")

    # Top KPIs
    if gap_data:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total Committed (USD)", fmt_usd(gap_data["total_committed_usd"]))
        c2.metric("Total Deployed (USD)", fmt_usd(gap_data["total_deployed_usd"]))
        c3.metric("Funding Gap", fmt_usd(gap_data["gap_usd"]))
        c4.metric("Coverage", f"{gap_data['coverage_percent']}%")

    col1, col2 = st.columns(2)

    with col1:
        if by_type_data:
            st.subheader("Resources by Type (USD)")
            type_df = pd.DataFrame(list(by_type_data.items()), columns=["Type", "Amount (USD)"])
            fig_h = px.bar(type_df.sort_values("Amount (USD)"),
                           x="Amount (USD)", y="Type", orientation="h",
                           color_discrete_sequence=[WHO_BLUE])
            fig_h.update_layout(height=280, margin=dict(t=10, b=0))
            st.plotly_chart(fig_h, use_container_width=True)

    with col2:
        if gap_data:
            st.subheader("Funding Gap Gauge")
            committed = gap_data["total_committed_usd"]
            needed = gap_data["estimated_need_usd"]
            fig_g = go.Figure(go.Indicator(
                mode="gauge+number",
                value=committed / 1_000_000,
                number={"suffix": "M USD"},
                gauge={
                    "axis": {"range": [0, needed / 1_000_000]},
                    "bar": {"color": WHO_BLUE},
                    "threshold": {
                        "line": {"color": ALERT_RED, "width": 3},
                        "value": needed / 1_000_000,
                    },
                },
                title={"text": f"Funding vs Need ({fmt_usd(needed)} total)"},
            ))
            fig_g.update_layout(height=260, margin=dict(t=50, b=0))
            st.plotly_chart(fig_g, use_container_width=True)

    # Funding timeline
    if not df.empty and "commitment_date" in df.columns:
        st.subheader("Funding Commitments Over Time")
        funding_df = df[df["resource_type"] == "Funding"].copy()
        if not funding_df.empty:
            funding_df["commitment_date"] = pd.to_datetime(funding_df["commitment_date"])
            funding_df = funding_df.sort_values("commitment_date")
            funding_df["cumulative"] = funding_df["amount"].cumsum()
            fig_line = px.line(funding_df, x="commitment_date", y="cumulative",
                               labels={"cumulative": "Cumulative USD", "commitment_date": "Date"},
                               markers=True, color_discrete_sequence=[WHO_BLUE])
            fig_line.update_layout(height=240, margin=dict(t=10, b=0))
            st.plotly_chart(fig_line, use_container_width=True)

    # Resources table with color coding
    st.subheader("All Resource Commitments")
    STATUS_COLORS = {
        "Deployed": "background-color:#d4edda",
        "Committed": "background-color:#fff3cd",
        "Pipeline": "background-color:#ffe0b2",
        "Under Review": "background-color:#f8d7da",
    }

    def color_status(val):
        return STATUS_COLORS.get(val, "")

    display_cols = ["partner_name", "resource_type", "description", "amount", "currency", "status", "next_report_due"]
    display_df = df[display_cols].copy() if all(c in df.columns for c in display_cols) else df
    st.dataframe(
        display_df.style.map(color_status, subset=["status"]) if "status" in display_df.columns else display_df,
        use_container_width=True, hide_index=True,
    )

    # Add resource form
    st.divider()
    st.subheader("Add Resource Commitment")
    partners_data, _ = api_get("partners", {"status": "Active"})
    if partners_data:
        partner_map = {p["name"]: p["id"] for p in partners_data}
        with st.form("add_resource_form"):
            g1, g2 = st.columns(2)
            r_partner = g1.selectbox("Partner *", list(partner_map.keys()))
            r_type = g2.selectbox("Resource Type *", ["Funding", "PPE", "Personnel", "Vaccines", "Logistics", "Technical"])
            g3, g4, g5 = st.columns(3)
            r_amount = g3.number_input("Amount", min_value=0.0, step=1000.0)
            r_currency = g4.selectbox("Currency", ["USD", "EUR", "GBP", "In-kind"])
            r_status = g5.selectbox("Status", ["Pipeline", "Committed", "Deployed", "Under Review"])
            r_desc = st.text_area("Description")
            g6, g7, g8 = st.columns(3)
            r_commit = g6.date_input("Commitment Date", value=date.today())
            r_deploy = g7.date_input("Deployment Date", value=None)
            r_freq = g8.selectbox("Reporting Frequency", ["Monthly", "Weekly", "Biweekly", "Quarterly", "Daily"])
            if st.form_submit_button("Add Resource", type="primary"):
                payload = {
                    "partner_id": partner_map[r_partner],
                    "resource_type": r_type,
                    "amount": r_amount,
                    "currency": r_currency,
                    "status": r_status,
                    "description": r_desc,
                    "commitment_date": r_commit.isoformat(),
                    "deployment_date": r_deploy.isoformat() if r_deploy else None,
                    "reporting_frequency": r_freq,
                }
                resp, code = api_post("resources", payload)
                if code in (200, 201):
                    st.success("Resource added!")
                    st.cache_data.clear()
                    st.rerun()
                else:
                    st.error(resp.get("message", "Error"))


# ══════════════════════════════════════════════════════════════════════════════
# PAGE 4 — 3W MATRIX
# ══════════════════════════════════════════════════════════════════════════════
elif page == "3W Matrix":
    st.title("3W Matrix — Who does What Where")

    matrix_data, err = api_get("activities/3w-matrix")
    if err:
        api_offline_msg(err)
        st.stop()

    if not matrix_data:
        st.warning("No activity data available.")
        st.stop()

    matrix = matrix_data["matrix"]
    locations = matrix_data["locations"]
    activity_types = matrix_data["activity_types"]

    # Build DataFrame for 3W table
    rows = []
    for loc in locations:
        row = {"Location": loc}
        for atype in activity_types:
            partners = matrix.get(loc, {}).get(atype, [])
            row[atype] = ", ".join(partners) if partners else ""
        rows.append(row)
    df_3w = pd.DataFrame(rows)

    st.subheader("Active Response Map (3W)")
    if df_3w.empty or "Location" not in df_3w.columns:
        st.info("No 3W data to display yet.")
    else:
        st.dataframe(df_3w.set_index("Location"), use_container_width=True)

    # Activity map (DRC provinces approximate coords)
    PROVINCE_COORDS = {
        "Equateur Province": {"lat": 0.5, "lon": 22.0},
        "Nord-Ubangi": {"lat": 3.5, "lon": 21.5},
        "Sud-Ubangi": {"lat": 2.5, "lon": 21.0},
        "Mbandaka": {"lat": 0.048, "lon": 18.26},
        "Bikoro": {"lat": -0.7, "lon": 18.1},
        "Ingende": {"lat": -0.26, "lon": 18.94},
        "Bolomba": {"lat": -0.4, "lon": 19.37},
    }

    activities, _ = api_get("activities")
    if activities:
        act_df = pd.DataFrame(activities)
        act_df["lat"] = act_df["location"].map(lambda x: PROVINCE_COORDS.get(x, {}).get("lat"))
        act_df["lon"] = act_df["location"].map(lambda x: PROVINCE_COORDS.get(x, {}).get("lon"))
        act_df_map = act_df.dropna(subset=["lat", "lon"])

        col1, col2 = st.columns([2, 1])
        with col1:
            st.subheader("Activity Locations (DRC)")
            if not act_df_map.empty:
                fig_map = px.scatter_mapbox(
                    act_df_map, lat="lat", lon="lon",
                    color="activity_type", hover_name="partner_name",
                    hover_data={"location": True, "status": True, "lat": False, "lon": False},
                    zoom=4, height=400,
                    mapbox_style="open-street-map",
                )
                fig_map.update_layout(margin=dict(t=0, b=0))
                st.plotly_chart(fig_map, use_container_width=True)

        with col2:
            st.subheader("By Activity Type")
            type_cnt = act_df["activity_type"].value_counts().reset_index()
            type_cnt.columns = ["Activity", "Count"]
            fig_act = px.bar(type_cnt, x="Count", y="Activity", orientation="h",
                             color_discrete_sequence=[WHO_BLUE])
            fig_act.update_layout(height=200, margin=dict(t=10, b=0))
            st.plotly_chart(fig_act, use_container_width=True)

            st.subheader("By Status")
            status_cnt = act_df["status"].value_counts().reset_index()
            status_cnt.columns = ["Status", "Count"]
            fig_st = px.pie(status_cnt, names="Status", values="Count", hole=0.4,
                            color_discrete_map={"Ongoing": SUCCESS_GREEN,
                                                "Planned": WHO_BLUE, "Completed": "#888"})
            fig_st.update_layout(height=200, margin=dict(t=10, b=0))
            st.plotly_chart(fig_st, use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE 5 — SITUATION REPORTS
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Situation Reports":
    st.title("Situation Reports")

    sitreps, err = api_get("sitreps")
    if err:
        api_offline_msg(err)
        st.stop()

    if sitreps:
        df_sr = pd.DataFrame(sitreps)
        df_sr["report_date"] = pd.to_datetime(df_sr["report_date"])
        df_sr = df_sr.sort_values("report_date")

        # Cases + deaths over time
        st.subheader("Epidemic Curve")
        fig_epi = go.Figure()
        fig_epi.add_trace(go.Scatter(
            x=df_sr["report_date"], y=df_sr["confirmed_cases"],
            name="Confirmed Cases", mode="lines+markers",
            line=dict(color=WARNING_ORANGE, width=2),
        ))
        fig_epi.add_trace(go.Scatter(
            x=df_sr["report_date"], y=df_sr["deaths"],
            name="Deaths", mode="lines+markers",
            line=dict(color=ALERT_RED, width=2),
        ))
        fig_epi.update_layout(height=300, margin=dict(t=10, b=0),
                               xaxis_title="Date", yaxis_title="Count",
                               legend=dict(orientation="h", yanchor="bottom", y=1.02))
        st.plotly_chart(fig_epi, use_container_width=True)

        # CFR trend
        col1, col2 = st.columns(2)
        with col1:
            st.subheader("Case Fatality Rate (%)")
            fig_cfr = px.line(df_sr, x="report_date", y="cfr_percent",
                              markers=True, color_discrete_sequence=[ALERT_RED])
            fig_cfr.update_layout(height=220, margin=dict(t=10, b=0))
            st.plotly_chart(fig_cfr, use_container_width=True)

        with col2:
            st.subheader("Funding Mobilized (USD)")
            fig_fund = px.bar(df_sr, x="report_date", y="total_funding_mobilized",
                              color_discrete_sequence=[WHO_BLUE])
            fig_fund.update_layout(height=220, margin=dict(t=10, b=0))
            st.plotly_chart(fig_fund, use_container_width=True)

        # Latest sitrep
        latest = sitreps[0]  # already sorted desc by API
        st.subheader(f"Latest SitRep — Day {latest['outbreak_day']} ({latest['report_date']})")
        lc1, lc2, lc3, lc4, lc5 = st.columns(5)
        lc1.metric("Confirmed Cases", latest["confirmed_cases"])
        lc2.metric("Deaths", latest["deaths"])
        lc3.metric("CFR", f"{latest['cfr_percent']}%")
        lc4.metric("HCW Affected", latest["healthcare_workers_affected"])
        lc5.metric("ETUs Operational", latest["etus_operational"])
        if latest.get("notes"):
            st.info(latest["notes"])

    # New sitrep form
    st.divider()
    st.subheader("Submit New Situation Report")
    with st.form("new_sitrep_form"):
        n1, n2 = st.columns(2)
        s_date = n1.date_input("Report Date", value=date.today())
        s_day = n2.number_input("Outbreak Day *", min_value=1, step=1, value=29)
        n3, n4 = st.columns(2)
        s_cases = n3.number_input("Confirmed Cases", min_value=0, step=1)
        s_deaths = n4.number_input("Deaths", min_value=0, step=1)
        n5, n6, n7 = st.columns(3)
        s_hcw = n5.number_input("HCW Affected", min_value=0, step=1)
        s_etu = n6.number_input("ETUs Operational", min_value=0, step=1)
        s_funding = n7.number_input("Funding Mobilized (USD)", min_value=0.0, step=100_000.0)
        s_gap = st.number_input("Funding Gap (USD)", min_value=0.0, step=100_000.0)
        s_notes = st.text_area("Notes / Key messages")
        s_author = st.text_input("Created By", value="WHO EOC")
        if st.form_submit_button("Submit SitRep", type="primary"):
            payload = {
                "report_date": s_date.isoformat(),
                "outbreak_day": int(s_day),
                "confirmed_cases": int(s_cases),
                "deaths": int(s_deaths),
                "healthcare_workers_affected": int(s_hcw),
                "etus_operational": int(s_etu),
                "total_funding_mobilized": float(s_funding),
                "funding_gap": float(s_gap),
                "notes": s_notes,
                "created_by": s_author,
            }
            resp, code = api_post("sitreps", payload)
            if code in (200, 201):
                st.success("Situation report submitted!")
                st.cache_data.clear()
                st.rerun()
            else:
                st.error(resp.get("message", "Error"))


# ══════════════════════════════════════════════════════════════════════════════
# PAGE 6 — REPORTS & EXPORT
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Reports & Export":
    st.title("Reports & Export")

    partners, _ = api_get("partners")
    resources, _ = api_get("resources")
    activities, _ = api_get("activities")
    sitreps, _ = api_get("sitreps")
    due_data, _ = api_get("resources/reporting-due", {"days": 14})

    # Reporting deadlines
    st.subheader("Upcoming Reporting Deadlines (14 days)")
    if due_data:
        overdue = due_data.get("overdue", [])
        due_soon = due_data.get("due_soon", [])
        if overdue:
            st.error(f"{len(overdue)} overdue reports")
            st.dataframe(pd.DataFrame(overdue)[["partner_name", "resource_type", "next_report_due", "reporting_frequency"]],
                         hide_index=True, use_container_width=True)
        if due_soon:
            st.warning(f"{len(due_soon)} reports due within 14 days")
            st.dataframe(pd.DataFrame(due_soon)[["partner_name", "resource_type", "next_report_due", "reporting_frequency"]],
                         hide_index=True, use_container_width=True)

    # Resource utilization
    if resources:
        st.subheader("Resource Utilization Summary")
        r_df = pd.DataFrame(resources)
        summary_tbl = r_df.groupby(["resource_type", "status"])["amount"].sum().reset_index()
        st.dataframe(summary_tbl, hide_index=True, use_container_width=True)

    st.divider()
    st.subheader("Export Data")

    col1, col2, col3, col4 = st.columns(4)

    # CSV exports
    if partners:
        p_df = pd.DataFrame(partners)
        col1.download_button(
            "Partners CSV", p_df.to_csv(index=False).encode(),
            file_name="evd_partners.csv", mime="text/csv",
        )
    if resources:
        r_df = pd.DataFrame(resources)
        col2.download_button(
            "Resources CSV", r_df.to_csv(index=False).encode(),
            file_name="evd_resources.csv", mime="text/csv",
        )
    if activities:
        a_df = pd.DataFrame(activities)
        col3.download_button(
            "Activities CSV", a_df.to_csv(index=False).encode(),
            file_name="evd_activities.csv", mime="text/csv",
        )
    if sitreps:
        s_df = pd.DataFrame(sitreps)
        col4.download_button(
            "SitReps CSV", s_df.to_csv(index=False).encode(),
            file_name="evd_sitreps.csv", mime="text/csv",
        )

    # Excel export
    st.divider()
    if st.button("Generate Full Excel Report"):
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            if partners:
                pd.DataFrame(partners).to_excel(writer, sheet_name="Partners", index=False)
            if resources:
                pd.DataFrame(resources).to_excel(writer, sheet_name="Resources", index=False)
            if activities:
                pd.DataFrame(activities).to_excel(writer, sheet_name="Activities", index=False)
            if sitreps:
                pd.DataFrame(sitreps).to_excel(writer, sheet_name="SitReps", index=False)
        buf.seek(0)
        st.download_button(
            "Download Excel Report",
            buf,
            file_name=f"EVD_Response_Report_{date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.success("Excel report ready for download.")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE 7 — GOVERNMENT ACTIVITIES
# ══════════════════════════════════════════════════════════════════════════════
elif page == "📋 Government Activities":
    # ── colour palette ────────────────────────────────────────────────────────
    C_BLUE   = "#0093D5"
    C_DARK   = "#003865"
    C_GREEN  = "#1E8449"
    C_AMBER  = "#D97706"
    C_RED    = "#C0392B"
    C_LBLUE  = "#EAF7FB"

    @st.cache_data(ttl=60)
    def _get_gov_summary():
        return api_get("government-activities/summary")

    @st.cache_data(ttl=60)
    def _get_gov_activities(pillar="", status="", priority=""):
        params = {}
        if pillar:   params["pillar"]   = pillar
        if status:   params["status"]   = status
        if priority: params["priority"] = priority
        return api_get("government-activities", params or None)

    PILLAR_SHORT = {
        1: "Leadership & Coord.",
        2: "Epi Surveillance",
        3: "Laboratory",
        4: "Case Mgmt/IPC",
        5: "RCCE",
        6: "Ops & Logistics",
        7: "Research & SI",
    }
    PILLAR_FULL = {
        1: "Leadership and Coordination",
        2: "Epidemiological Surveillance",
        3: "Laboratory and Diagnostics",
        4: "Case Management, IPC/WASH and SDB",
        5: "Risk Communication and Community Engagement",
        6: "Operational Support and Logistics",
        7: "Research and Strategic Information",
    }

    st.title("Government Activity Register")
    st.caption("Source of truth for all preparedness activities across 7 pillars")

    col_refresh, _ = st.columns([1, 6])
    with col_refresh:
        if st.button("🔄 Refresh", key="ga_refresh"):
            st.cache_data.clear()
            st.rerun()

    summary, err = _get_gov_summary()
    if err:
        st.warning(f"API offline — {err}")
        st.info("Start Flask: `python app.py`")
        st.stop()

    # ── Top metrics ───────────────────────────────────────────────────────────
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Total Activities", summary.get("total_activities", 0))
    m2.metric("Govt Budget (USD)",
              f"${summary.get('total_cost_usd', 0):,.0f}")
    m3.metric("Partner Support (USD)",
              f"${summary.get('total_partner_support', 0):,.0f}")
    m4.metric("Funding Gap (USD)",
              f"${summary.get('total_funding_gap', 0):,.0f}")

    # ── Coverage progress bar ─────────────────────────────────────────────────
    cov = summary.get("coverage_pct", 0)
    bar_color = C_GREEN if cov >= 75 else (C_AMBER if cov >= 50 else C_RED)
    st.markdown(
        f"**Overall Funding Coverage**  &nbsp; "
        f"<span style='color:{bar_color};font-weight:700;'>{cov:.1f}%</span> "
        f"of total budget covered by partners",
        unsafe_allow_html=True,
    )
    st.progress(min(int(cov), 100))

    st.divider()

    # ── Filters ───────────────────────────────────────────────────────────────
    fc1, fc2, fc3, fc4 = st.columns(4)
    with fc1:
        f_pillar = st.selectbox(
            "Pillar", ["All"] + [f"Pillar {n} — {v}" for n, v in PILLAR_FULL.items()],
            key="ga_f_pillar",
        )
    with fc2:
        f_status = st.selectbox(
            "Status", ["All", "Planned", "Active", "Completed"],
            key="ga_f_status",
        )
    with fc3:
        f_priority = st.selectbox(
            "Priority", ["All", "High", "Medium", "Low"],
            key="ga_f_priority",
        )
    with fc4:
        f_coverage = st.selectbox(
            "Coverage", ["All", "Fully funded", "Partially funded", "Unfunded"],
            key="ga_f_coverage",
        )

    pillar_q   = f_pillar.split("—")[1].strip() if f_pillar != "All" else ""
    status_q   = f_status   if f_status   != "All" else ""
    priority_q = f_priority if f_priority != "All" else ""

    act_data, err2 = _get_gov_activities(pillar_q, status_q, priority_q)
    if err2 or not act_data:
        st.warning("No activity data available.")
        st.stop()

    grouped = act_data.get("grouped", [])

    # Apply coverage filter client-side
    if f_coverage != "All":
        for grp in grouped:
            if f_coverage == "Fully funded":
                grp["activities"] = [a for a in grp["activities"] if a["funding_gap"] == 0 and a["total_cost_usd"] > 0]
            elif f_coverage == "Partially funded":
                grp["activities"] = [a for a in grp["activities"] if 0 < a["funding_gap"] < a["total_cost_usd"]]
            elif f_coverage == "Unfunded":
                grp["activities"] = [a for a in grp["activities"] if a["total_partner_support"] == 0 and a["total_cost_usd"] > 0]
        grouped = [g for g in grouped if g["activities"]]

    # ── Init session state for edit ───────────────────────────────────────────
    if "ga_edit_id" not in st.session_state:
        st.session_state.ga_edit_id = None

    # ── Pillar accordions ─────────────────────────────────────────────────────
    for grp in sorted(grouped, key=lambda x: x.get("pillar_number", 99)):
        pnum   = grp.get("pillar_number", 0)
        pname  = grp.get("pillar_name", "")
        acts   = grp.get("activities", [])

        if not acts:
            continue

        p_cost    = sum(a["total_cost_usd"] for a in acts)
        p_support = sum(a["total_partner_support"] for a in acts)
        p_gap     = sum(a["funding_gap"] for a in acts)
        p_cov     = round(p_support / p_cost * 100, 1) if p_cost else 0

        title_color = C_GREEN if p_cov >= 75 else (C_AMBER if p_cov >= 50 else C_RED)
        expander_label = (
            f"Pillar {pnum} — {pname}  |  "
            f"${p_cost:,.0f}  |  Coverage: {p_cov:.0f}%  |  Gap: ${p_gap:,.0f}"
        )

        with st.expander(expander_label, expanded=(pnum == 1)):
            st.markdown(
                f"<div style='color:{title_color};font-weight:700;margin-bottom:8px;'>"
                f"Pillar {pnum}: {p_cov:.1f}% funded | "
                f"${p_support:,.0f} of ${p_cost:,.0f} covered</div>",
                unsafe_allow_html=True,
            )

            # Table header
            hcols = st.columns([0.4, 3.5, 1.2, 1.2, 1.2, 0.8, 0.9, 0.8, 0.7])
            headers = ["#", "Activity", "Cost (USD)", "Support (USD)", "Gap (USD)",
                       "Cov %", "Status", "Priority", ""]
            for hc, ht in zip(hcols, headers):
                hc.markdown(f"**{ht}**")
            st.divider()

            for a in acts:
                gap_color = C_GREEN if a["funding_gap"] == 0 else \
                            (C_AMBER if a["funding_gap"] < a["total_cost_usd"] * 0.5 else C_RED)
                cov_color = C_GREEN if a["coverage_pct"] >= 75 else \
                            (C_AMBER if a["coverage_pct"] >= 50 else C_RED)

                row = st.columns([0.4, 3.5, 1.2, 1.2, 1.2, 0.8, 0.9, 0.8, 0.7])
                row[0].write(a.get("activity_number", ""))
                row[1].write(a["activity_name"])
                row[2].write(f"${a['total_cost_usd']:,.0f}")
                row[3].write(f"${a['total_partner_support']:,.0f}")
                row[4].markdown(
                    f"<span style='color:{gap_color};font-weight:600;'>"
                    f"${a['funding_gap']:,.0f}</span>",
                    unsafe_allow_html=True,
                )
                row[5].markdown(
                    f"<span style='color:{cov_color};font-weight:600;'>"
                    f"{a['coverage_pct']:.0f}%</span>",
                    unsafe_allow_html=True,
                )
                row[6].write(a.get("status", ""))
                row[7].write(a.get("priority", ""))
                if row[8].button("✏️", key=f"edit_{a['id']}", help="Edit activity"):
                    st.session_state.ga_edit_id = a["id"]

    # ── Edit form ─────────────────────────────────────────────────────────────
    if st.session_state.ga_edit_id is not None:
        edit_id = st.session_state.ga_edit_id
        act_detail, _ = api_get(f"government-activities/{edit_id}")
        if act_detail:
            st.divider()
            st.subheader(f"Edit Activity #{act_detail.get('activity_number', edit_id)}")
            with st.form(f"edit_activity_{edit_id}"):
                e1, e2 = st.columns(2)
                e_pillar = e1.selectbox(
                    "Pillar",
                    options=list(PILLAR_FULL.values()),
                    index=max(0, (act_detail.get("pillar_number") or 1) - 1),
                    key=f"ep_{edit_id}",
                )
                e_sub = e2.text_input("Sub-section",
                                      value=act_detail.get("sub_section", ""),
                                      key=f"es_{edit_id}")
                e3, e4 = st.columns([1, 3])
                e_num = e3.text_input("Activity Number",
                                      value=act_detail.get("activity_number", ""),
                                      key=f"en_{edit_id}")
                e_name = e4.text_area("Activity Name",
                                      value=act_detail.get("activity_name", ""),
                                      height=80, key=f"ename_{edit_id}")
                e5, e6, e7 = st.columns(3)
                e_cost = e5.number_input("Total Cost USD",
                                         value=float(act_detail.get("total_cost_usd", 0)),
                                         min_value=0.0, step=1000.0,
                                         key=f"ec_{edit_id}")
                e_status = e6.selectbox("Status",
                                        ["Planned", "Active", "Completed", "Suspended"],
                                        index=["Planned", "Active", "Completed", "Suspended"]
                                        .index(act_detail.get("status", "Planned")),
                                        key=f"est_{edit_id}")
                e_priority = e7.selectbox("Priority", ["High", "Medium", "Low"],
                                          index=["High", "Medium", "Low"]
                                          .index(act_detail.get("priority", "Medium")),
                                          key=f"epri_{edit_id}")
                e_notes = st.text_area("Notes",
                                       value=act_detail.get("notes", "") or "",
                                       height=60, key=f"enotes_{edit_id}")

                save_col, susp_col, cancel_col = st.columns(3)
                save_btn   = save_col.form_submit_button("💾 Save Changes", type="primary")
                susp_btn   = susp_col.form_submit_button("🚫 Mark Suspended")
                cancel_btn = cancel_col.form_submit_button("✖ Cancel")

                pnum_map = {v: k for k, v in PILLAR_FULL.items()}

                if save_btn:
                    payload = {
                        "activity_number": e_num,
                        "pillar": e_pillar,
                        "pillar_number": pnum_map.get(e_pillar, act_detail.get("pillar_number")),
                        "sub_section": e_sub,
                        "activity_name": e_name,
                        "total_cost_usd": e_cost,
                        "status": e_status,
                        "priority": e_priority,
                        "notes": e_notes,
                    }
                    resp, code = api_put(f"government-activities/{edit_id}", payload)
                    if code == 200:
                        st.success("Activity updated!")
                        st.session_state.ga_edit_id = None
                        st.cache_data.clear()
                        st.rerun()
                    else:
                        st.error(resp.get("message", "Update failed"))

                if susp_btn:
                    resp, code = api_delete(f"government-activities/{edit_id}")
                    if code == 200:
                        st.warning("Activity suspended.")
                        st.session_state.ga_edit_id = None
                        st.cache_data.clear()
                        st.rerun()
                    else:
                        st.error(resp.get("message", "Suspension failed"))

                if cancel_btn:
                    st.session_state.ga_edit_id = None
                    st.rerun()

    # ── Add new activity ──────────────────────────────────────────────────────
    st.divider()
    with st.expander("➕ Add New Activity", expanded=False):
        with st.form("add_gov_activity"):
            na1, na2 = st.columns(2)
            na_pillar = na1.selectbox("Pillar *",
                                      [f"Pillar {n} — {v}" for n, v in PILLAR_FULL.items()])
            na_sub    = na2.text_input("Sub-section (e.g. 'A. Leadership and Coordination')")
            na3, na4 = st.columns([1, 3])
            na_num    = na3.text_input("Activity Number (e.g. '16.0')")
            na_name   = na4.text_area("Activity Name *", height=80)
            na5, na6, na7 = st.columns(3)
            na_cost     = na5.number_input("Total Cost USD", min_value=0.0, step=1000.0)
            na_status   = na6.selectbox("Status", ["Planned", "Active", "Completed", "Suspended"])
            na_priority = na7.selectbox("Priority", ["High", "Medium", "Low"])
            na_notes    = st.text_area("Notes (optional)", height=60)

            if st.form_submit_button("➕ Add Activity", type="primary"):
                if not na_name.strip():
                    st.error("Activity name is required.")
                else:
                    pnum_sel = int(na_pillar.split(" ")[1])
                    resp, code = api_post("government-activities", {
                        "activity_number": na_num,
                        "pillar": PILLAR_FULL[pnum_sel],
                        "pillar_number": pnum_sel,
                        "sub_section": na_sub,
                        "activity_name": na_name,
                        "total_cost_usd": na_cost,
                        "status": na_status,
                        "priority": na_priority,
                        "notes": na_notes,
                    })
                    if code in (200, 201):
                        st.success("Activity added!")
                        st.cache_data.clear()
                        st.rerun()
                    else:
                        st.error(resp.get("message", "Error adding activity"))

    # ── Pillar summary chart ──────────────────────────────────────────────────
    st.divider()
    st.subheader("Pillar Funding Overview")
    by_pillar = summary.get("by_pillar", [])
    if by_pillar:
        chart_df = pd.DataFrame(by_pillar)
        chart_df["short_name"] = chart_df["pillar_number"].map(PILLAR_SHORT)

        fig = go.Figure()
        fig.add_trace(go.Bar(
            name="Govt Budget",
            x=chart_df["short_name"],
            y=chart_df["total_cost"],
            marker_color="#888888",
            opacity=0.7,
        ))
        fig.add_trace(go.Bar(
            name="Partner Support",
            x=chart_df["short_name"],
            y=chart_df["total_support"],
            marker_color=C_BLUE,
        ))
        fig.add_trace(go.Scatter(
            name="Coverage %",
            x=chart_df["short_name"],
            y=chart_df["coverage_pct"],
            mode="lines+markers",
            yaxis="y2",
            line=dict(color=C_GREEN, width=2),
            marker=dict(size=8),
        ))
        fig.update_layout(
            barmode="overlay",
            height=380,
            margin=dict(t=20, b=0),
            legend=dict(orientation="h", yanchor="bottom", y=1.02),
            yaxis=dict(title="USD"),
            yaxis2=dict(title="Coverage %", overlaying="y", side="right",
                        range=[0, 110]),
        )
        st.plotly_chart(fig, use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE 8 — PARTNER CONTRIBUTIONS
# ══════════════════════════════════════════════════════════════════════════════
elif page == "🤝 Partner Contributions":
    C_BLUE  = "#0093D5"
    C_DARK  = "#003865"
    C_GREEN = "#1E8449"
    C_AMBER = "#D97706"
    C_RED   = "#C0392B"
    C_LBLUE = "#EAF7FB"

    @st.cache_data(ttl=60)
    def _get_matrix():
        return api_get("partner-contributions/matrix")

    @st.cache_data(ttl=60)
    def _get_gaps():
        return api_get("partner-contributions/gaps")

    @st.cache_data(ttl=60)
    def _get_all_partners():
        return api_get("partners")

    @st.cache_data(ttl=60)
    def _get_gov_flat():
        d, e = api_get("government-activities")
        if e or not d:
            return None, e
        return d.get("flat", []), e

    st.title("Partner Contributions")
    st.caption("Track partner pledges across all 7 preparedness pillars")

    col_refresh, _ = st.columns([1, 6])
    with col_refresh:
        if st.button("🔄 Refresh", key="pc_refresh"):
            st.cache_data.clear()
            st.rerun()

    # ── Top summary metrics ───────────────────────────────────────────────────
    matrix_data, err_m = _get_matrix()
    if err_m:
        st.warning(f"API offline — {err_m}")
        st.info("Start Flask: `python app.py`")
        st.stop()

    all_contribs, _ = api_get("partner-contributions")
    all_contribs = all_contribs or []

    total_pledged    = sum(c.get("amount_pledged_usd", 0) for c in all_contribs)
    total_disbursed  = sum(c.get("amount_disbursed_usd", 0) for c in all_contribs)
    total_balance    = sum(c.get("balance_usd", 0) for c in all_contribs)
    num_partners_con = len({c.get("partner_id") for c in all_contribs})

    tm1, tm2, tm3, tm4 = st.columns(4)
    tm1.metric("Partners Contributing", num_partners_con)
    tm2.metric("Total Pledged (USD)",    f"${total_pledged:,.0f}")
    tm3.metric("Total Disbursed (USD)",  f"${total_disbursed:,.0f}")
    tm4.metric("Balance Remaining (USD)", f"${total_balance:,.0f}")

    st.divider()

    # ── View selector ─────────────────────────────────────────────────────────
    pc_view = st.radio(
        "View",
        ["📊 Matrix View", "📋 By Activity", "🏢 By Partner", "⚠️ Gaps"],
        horizontal=True,
        key="pc_view_radio",
        label_visibility="collapsed",
    )

    # ══════════════════════════════════════════════════════════════════════════
    # VIEW 1 — MATRIX VIEW
    # ══════════════════════════════════════════════════════════════════════════
    if pc_view == "📊 Matrix View":
        if not matrix_data:
            st.info("No contribution data yet.")
            st.stop()

        activities   = matrix_data.get("activities", [])
        partners     = matrix_data.get("partners", [])
        matrix       = matrix_data.get("matrix", {})
        row_totals   = matrix_data.get("row_totals", {})
        col_totals   = matrix_data.get("col_totals", {})
        grand_total  = matrix_data.get("grand_total", 0)

        # ── Filters ───────────────────────────────────────────────────────────
        mf1, mf2, mf3 = st.columns(3)
        with mf1:
            PILLAR_FULL_OPTS = {
                1: "Leadership and Coordination",
                2: "Epidemiological Surveillance",
                3: "Laboratory and Diagnostics",
                4: "Case Management, IPC/WASH and SDB",
                5: "Risk Communication and Community Engagement",
                6: "Operational Support and Logistics",
                7: "Research and Strategic Information",
            }
            f_pil = st.selectbox("Filter by Pillar",
                                 ["All Pillars"] + list(PILLAR_FULL_OPTS.values()),
                                 key="mx_fpil")
        with mf2:
            partner_names_all = [p["name"] for p in partners]
            f_part = st.multiselect("Show Partners", partner_names_all,
                                    default=partner_names_all, key="mx_fpart")
        with mf3:
            show_gaps_only = st.toggle("Show only activities with gaps", key="mx_gaps")

        # Filter activities
        vis_activities = activities
        if f_pil != "All Pillars":
            vis_activities = [a for a in activities if a.get("pillar") == f_pil]
        if show_gaps_only:
            vis_activities = [a for a in vis_activities if a.get("funding_gap", 0) > 0]

        # Filter partners
        vis_partners = [p for p in partners if p["name"] in f_part]

        # ── Build matrix dataframe ─────────────────────────────────────────────
        if not vis_activities:
            st.info("No activities match the current filters.")
        else:
            rows = []
            prev_pillar = None
            for a in vis_activities:
                if a.get("pillar") != prev_pillar:
                    prev_pillar = a.get("pillar")
                    rows.append({"__type": "header", "Activity": f"▌ {prev_pillar}"})
                row = {
                    "__type": "activity",
                    "Activity": f"{a.get('activity_number','')}  {a['activity_name'][:65]}",
                    "Govt Cost": a.get("total_cost_usd", 0),
                }
                a_id = str(a["id"])
                for p in vis_partners:
                    p_id = str(p["id"])
                    cell = matrix.get(a_id, {}).get(p_id)
                    row[p["name"]] = cell["amount_pledged"] if cell else 0
                row["Total Support"] = float(row_totals.get(a_id, 0))
                row["Funding Gap"]   = max(0, a.get("total_cost_usd", 0) - row["Total Support"])
                rows.append(row)

            df_matrix = pd.DataFrame(rows)
            df_matrix = df_matrix.fillna(0)

            # Format currency columns for display
            money_cols = ["Govt Cost"] + [p["name"] for p in vis_partners] + ["Total Support", "Funding Gap"]
            money_cols = [c for c in money_cols if c in df_matrix.columns]

            def fmt_cell(v):
                if isinstance(v, (int, float)) and v > 0:
                    return f"${v:,.0f}"
                if isinstance(v, (int, float)) and v == 0:
                    return "—"
                return v

            display_df = df_matrix.copy()
            for col in money_cols:
                if col in display_df.columns:
                    display_df[col] = display_df[col].apply(
                        lambda v: fmt_cell(v) if display_df.loc[
                            display_df[col] == v, "__type"
                        ].values[0] != "header" if len(
                            display_df[display_df[col] == v]
                        ) > 0 else fmt_cell(v) else v
                    )

            # Simpler formatting without the problematic lambda
            for col in money_cols:
                if col in df_matrix.columns:
                    display_df[col] = df_matrix.apply(
                        lambda row: ("" if row["__type"] == "header"
                                     else (f"${row[col]:,.0f}" if row[col] > 0 else "—")),
                        axis=1,
                    )

            display_df = display_df.drop(columns=["__type"])

            # Colour-code using pandas Styler
            def style_gap(val):
                if val == "" or val == "—":
                    return ""
                try:
                    v = float(str(val).replace("$", "").replace(",", ""))
                except ValueError:
                    return ""
                if v == 0:
                    return f"color:{C_GREEN};font-weight:600"
                return f"color:{C_RED};font-weight:600"

            styled = display_df.style
            if "Funding Gap" in display_df.columns:
                styled = styled.applymap(style_gap, subset=["Funding Gap"])

            st.dataframe(styled, use_container_width=True, hide_index=True,
                         height=min(35 * len(display_df) + 38, 700))

            # Grand total row
            gt_cols  = st.columns([3, 1] + [1] * len(vis_partners) + [1, 1])
            gt_vals  = ["**GRAND TOTAL**", f"**${sum(a.get('total_cost_usd',0) for a in vis_activities):,.0f}**"]
            for p in vis_partners:
                gt_vals.append(f"**${float(col_totals.get(str(p['id']),0)):,.0f}**")
            gt_vals += [f"**${grand_total:,.0f}**", ""]
            for gc, gv in zip(gt_cols, gt_vals):
                gc.markdown(gv)

            # ── Export to Excel ───────────────────────────────────────────────
            st.divider()
            if st.button("📥 Export Matrix to Excel", key="mx_export"):
                buf = io.BytesIO()
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                wb = Workbook()
                ws = wb.active
                ws.title = "Partner Support Matrix"

                header_fill  = PatternFill("solid", fgColor="003865")
                pillar_fill  = PatternFill("solid", fgColor="003865")
                support_fill = PatternFill("solid", fgColor="0093D5")
                gap_red      = PatternFill("solid", fgColor="F1948A")
                gap_green    = PatternFill("solid", fgColor="A9DFBF")
                white_font   = Font(color="FFFFFF", bold=True)
                bold_font    = Font(bold=True)

                # Column headers
                cols_hdr = (["Activity", "Govt Cost (USD)"]
                            + [p["name"] for p in vis_partners]
                            + ["Total Support (USD)", "Funding Gap (USD)"])
                ws.append(cols_hdr)
                for cell in ws[1]:
                    cell.fill  = header_fill
                    cell.font  = white_font
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)

                prev_pillar_xl = None
                for a in vis_activities:
                    if a.get("pillar") != prev_pillar_xl:
                        prev_pillar_xl = a.get("pillar")
                        ws.append([f"▌ {prev_pillar_xl}"] + [""] * (len(cols_hdr) - 1))
                        for cell in ws[ws.max_row]:
                            cell.fill = pillar_fill
                            cell.font = white_font

                    a_id     = str(a["id"])
                    row_data = [
                        f"{a.get('activity_number','')} {a['activity_name']}",
                        a.get("total_cost_usd", 0),
                    ]
                    for p in vis_partners:
                        p_id = str(p["id"])
                        cell_val = matrix.get(a_id, {}).get(p_id)
                        row_data.append(cell_val["amount_pledged"] if cell_val else 0)
                    t_support = float(row_totals.get(a_id, 0))
                    gap_val   = max(0, a.get("total_cost_usd", 0) - t_support)
                    row_data += [t_support, gap_val]
                    ws.append(row_data)

                    # Colour gap cell
                    gap_cell = ws.cell(row=ws.max_row, column=len(cols_hdr))
                    gap_cell.fill = gap_green if gap_val == 0 else gap_red

                    # Colour support cell
                    ws.cell(row=ws.max_row,
                            column=len(cols_hdr) - 1).fill = support_fill

                # Column widths
                ws.column_dimensions["A"].width = 60
                for i in range(2, len(cols_hdr) + 1):
                    ws.column_dimensions[get_column_letter(i)].width = 18

                wb.save(buf)
                buf.seek(0)
                st.download_button(
                    "⬇ Download Partner Support Matrix",
                    buf,
                    file_name=f"Partner_Support_Matrix_{date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                st.success("Excel export ready.")

    # ══════════════════════════════════════════════════════════════════════════
    # VIEW 2 — BY ACTIVITY
    # ══════════════════════════════════════════════════════════════════════════
    elif pc_view == "📋 By Activity":
        flat_acts, err_f = _get_gov_flat()
        all_partners, _  = _get_all_partners()

        if err_f or not flat_acts:
            st.warning("Could not load activities.")
            st.stop()

        if "pc_sel_act" not in st.session_state:
            st.session_state.pc_sel_act = None

        col_left, col_right = st.columns([1.2, 2.8])

        with col_left:
            st.subheader("Activities")
            PILLAR_FULL_OPTS2 = {
                1: "Leadership and Coordination",
                2: "Epidemiological Surveillance",
                3: "Laboratory and Diagnostics",
                4: "Case Management, IPC/WASH and SDB",
                5: "Risk Communication and Community Engagement",
                6: "Operational Support and Logistics",
                7: "Research and Strategic Information",
            }
            prev_pil = None
            for a in sorted(flat_acts, key=lambda x: (x.get("pillar_number", 99),
                                                        x.get("activity_number", ""))):
                if a.get("pillar") != prev_pil:
                    prev_pil = a.get("pillar")
                    st.markdown(
                        f"<div style='background:{C_DARK};color:white;"
                        f"padding:4px 8px;border-radius:4px;margin:6px 0 2px;font-size:.8rem;"
                        f"font-weight:700;'>{prev_pil}</div>",
                        unsafe_allow_html=True,
                    )
                label = f"{a.get('activity_number','')} {a['activity_name'][:50]}..."
                if st.button(label, key=f"act_sel_{a['id']}", use_container_width=True):
                    st.session_state.pc_sel_act = a["id"]

        with col_right:
            if st.session_state.pc_sel_act:
                act_id = st.session_state.pc_sel_act
                act_detail, _ = api_get(f"partner-contributions/by-activity/{act_id}")
                if act_detail:
                    activity = act_detail.get("activity", {})
                    contribs = act_detail.get("contributions", [])

                    st.subheader(activity.get("activity_name", ""))
                    d1, d2, d3 = st.columns(3)
                    d1.metric("Total Cost", f"${activity.get('total_cost_usd', 0):,.0f}")
                    d2.metric("Partner Support",
                              f"${activity.get('total_partner_support', 0):,.0f}")
                    d3.metric("Funding Gap",
                              f"${activity.get('funding_gap', 0):,.0f}")

                    st.caption(
                        f"Pillar: {activity.get('pillar')}  |  "
                        f"Status: {activity.get('status')}  |  "
                        f"Priority: {activity.get('priority')}"
                    )

                    cov_pct = activity.get("coverage_pct", 0)
                    cov_color = C_GREEN if cov_pct >= 75 else (C_AMBER if cov_pct >= 50 else C_RED)
                    st.markdown(
                        f"<span style='color:{cov_color};font-weight:700;'>"
                        f"{cov_pct:.0f}%</span> funded",
                        unsafe_allow_html=True,
                    )
                    st.progress(min(int(cov_pct), 100))

                    if contribs:
                        st.markdown("**Partner Contributions:**")
                        contrib_df = pd.DataFrame(contribs)[[
                            "partner_name", "amount_pledged_usd",
                            "amount_available_usd", "amount_to_mobilize_usd",
                            "amount_disbursed_usd", "balance_usd",
                            "modality", "status",
                        ]]
                        contrib_df.columns = [
                            "Partner", "Pledged (USD)", "Available (USD)",
                            "To Mobilize (USD)", "Disbursed (USD)",
                            "Balance (USD)", "Modality", "Status",
                        ]
                        st.dataframe(contrib_df, hide_index=True,
                                     use_container_width=True)
                    else:
                        st.info("No partner contributions recorded yet for this activity.")

                    # Add contribution form
                    st.divider()
                    with st.expander("➕ Add Partner Contribution", expanded=False):
                        partner_map = {p["name"]: p["id"]
                                       for p in (all_partners or [])}
                        with st.form(f"add_contrib_{act_id}"):
                            ac1, ac2 = st.columns(2)
                            ac_partner = ac1.selectbox("Partner *",
                                                       list(partner_map.keys()),
                                                       key=f"acp_{act_id}")
                            ac_modality = ac2.selectbox(
                                "Modality",
                                ["Direct Implementation", "Transfer to Government",
                                 "In-kind", "Co-implementation"],
                                key=f"acm_{act_id}",
                            )
                            ac3, ac4 = st.columns(2)
                            ac_pledged  = ac3.number_input("Amount Pledged (USD)",
                                                           min_value=0.0, step=1000.0,
                                                           key=f"acpl_{act_id}")
                            ac_avail    = ac4.number_input("Amount Available (USD)",
                                                           min_value=0.0, step=1000.0,
                                                           key=f"acav_{act_id}")
                            ac5, ac6 = st.columns(2)
                            ac_mob  = ac5.number_input("To Mobilize (USD)",
                                                       min_value=0.0, step=1000.0,
                                                       key=f"acmob_{act_id}")
                            ac_disb = ac6.number_input("Disbursed (USD)",
                                                        min_value=0.0, step=1000.0,
                                                        key=f"acd_{act_id}")
                            ac_status = st.selectbox(
                                "Status",
                                ["Pledged", "Confirmed", "Disbursed", "Cancelled"],
                                key=f"acs_{act_id}",
                            )
                            ac_notes = st.text_area("Notes", height=60,
                                                    key=f"acn_{act_id}")
                            if st.form_submit_button("💾 Save Contribution",
                                                     type="primary"):
                                resp, code = api_post("partner-contributions", {
                                    "government_activity_id": act_id,
                                    "partner_id": partner_map[ac_partner],
                                    "amount_pledged_usd":     ac_pledged,
                                    "amount_available_usd":   ac_avail,
                                    "amount_to_mobilize_usd": ac_mob,
                                    "amount_disbursed_usd":   ac_disb,
                                    "modality": ac_modality,
                                    "status":   ac_status,
                                    "notes":    ac_notes,
                                })
                                if code in (200, 201):
                                    st.success("Contribution saved!")
                                    st.cache_data.clear()
                                    st.rerun()
                                else:
                                    st.error(resp.get("message", "Error"))
            else:
                st.info("← Select an activity from the left to see details.")

    # ══════════════════════════════════════════════════════════════════════════
    # VIEW 3 — BY PARTNER
    # ══════════════════════════════════════════════════════════════════════════
    elif pc_view == "🏢 By Partner":
        all_partners, _ = _get_all_partners()

        # Only show partners that have contributions
        contrib_partner_ids = {c.get("partner_id") for c in all_contribs}
        active_partners = [p for p in (all_partners or [])
                           if p["id"] in contrib_partner_ids]

        if not active_partners:
            st.info("No partner contributions recorded yet.")
            st.stop()

        col_left2, col_right2 = st.columns([1, 3])

        with col_left2:
            st.subheader("Partners")
            for p in sorted(active_partners, key=lambda x: x["name"]):
                p_total = sum(c.get("amount_pledged_usd", 0)
                              for c in all_contribs if c.get("partner_id") == p["id"])
                p_acts  = sum(1 for c in all_contribs if c.get("partner_id") == p["id"])
                if st.button(
                    f"{p['name']}\n${p_total:,.0f} | {p_acts} activities",
                    key=f"psel_{p['id']}", use_container_width=True,
                ):
                    st.session_state["pc_sel_partner"] = p["id"]

        with col_right2:
            sel_pid = st.session_state.get("pc_sel_partner")
            if sel_pid:
                pdet, _ = api_get(f"partner-contributions/by-partner/{sel_pid}")
                if pdet:
                    p_info = pdet.get("partner", {})
                    st.subheader(p_info.get("name", ""))
                    st.caption(f"Type: {p_info.get('type', '')} | "
                               f"Activities supported: {pdet.get('activity_count', 0)}")

                    pm1, pm2, pm3, pm4 = st.columns(4)
                    pm1.metric("Activities",
                               pdet.get("activity_count", 0))
                    pm2.metric("Total Pledged",
                               f"${pdet.get('total_pledged', 0):,.0f}")
                    pm3.metric("Total Disbursed",
                               f"${pdet.get('total_disbursed', 0):,.0f}")
                    pm4.metric("Balance",
                               f"${pdet.get('total_balance', 0):,.0f}")

                    # Donut chart by pillar
                    by_pillar = pdet.get("by_pillar", {})
                    if by_pillar:
                        fig_donut = px.pie(
                            names=list(by_pillar.keys()),
                            values=list(by_pillar.values()),
                            hole=0.45,
                            title="Pledges by Pillar",
                            color_discrete_sequence=px.colors.sequential.Blues_r,
                        )
                        fig_donut.update_layout(height=320,
                                                margin=dict(t=40, b=0))
                        st.plotly_chart(fig_donut, use_container_width=True)

                    # Contributions table
                    contribs_p = pdet.get("contributions", [])
                    if contribs_p:
                        p_df = pd.DataFrame(contribs_p)[[
                            "pillar", "activity_name", "amount_pledged_usd",
                            "amount_available_usd", "amount_disbursed_usd",
                            "balance_usd", "modality",
                        ]]
                        p_df.columns = [
                            "Pillar", "Activity", "Pledged (USD)", "Available (USD)",
                            "Disbursed (USD)", "Balance (USD)", "Modality",
                        ]
                        st.dataframe(p_df, hide_index=True, use_container_width=True)
            else:
                st.info("← Select a partner from the left.")

    # ══════════════════════════════════════════════════════════════════════════
    # VIEW 4 — GAPS ANALYSIS
    # ══════════════════════════════════════════════════════════════════════════
    elif pc_view == "⚠️ Gaps":
        gaps, err_g = _get_gaps()
        if err_g or not gaps:
            st.info("No gap data available. Run seed_data.py to populate activities.")
            st.stop()

        st.subheader("Critical Gaps Leaderboard")
        st.caption("Activities sorted by funding gap (largest first). "
                   "Red = critical (<25% funded), Amber = at risk, Green = adequate (≥75%).")

        def row_color(cov):
            if cov < 25:
                return "background-color:#FADBD8"
            if cov < 75:
                return "background-color:#FDEBD0"
            return "background-color:#D5F5E3"

        gaps_df = pd.DataFrame(gaps)[[
            "pillar", "activity_name", "total_cost",
            "total_support", "gap", "coverage_pct",
            "status", "priority",
        ]]
        gaps_df.columns = [
            "Pillar", "Activity", "Govt Cost (USD)",
            "Partner Support (USD)", "Gap (USD)", "Coverage %",
            "Status", "Priority",
        ]
        for col in ["Govt Cost (USD)", "Partner Support (USD)", "Gap (USD)"]:
            gaps_df[col] = gaps_df[col].apply(lambda v: f"${v:,.0f}")

        styled_gaps = gaps_df.style.apply(
            lambda row: [row_color(row["Coverage %"])] * len(row), axis=1
        )
        st.dataframe(styled_gaps, use_container_width=True, hide_index=True,
                     height=min(35 * len(gaps_df) + 38, 600))

        # Gap by pillar bar chart
        st.divider()
        st.subheader("Funding Gap by Pillar")
        pillar_summary, _ = api_get("government-activities/summary")
        if pillar_summary:
            bp = pillar_summary.get("by_pillar", [])
            if bp:
                bp_df = pd.DataFrame(bp)
                fig_bars = go.Figure()
                fig_bars.add_trace(go.Bar(
                    name="Funded",
                    y=[f"Pillar {r['pillar_number']}" for _, r in bp_df.iterrows()],
                    x=bp_df["total_support"],
                    orientation="h",
                    marker_color=C_GREEN,
                ))
                fig_bars.add_trace(go.Bar(
                    name="Gap",
                    y=[f"Pillar {r['pillar_number']}" for _, r in bp_df.iterrows()],
                    x=bp_df["gap"],
                    orientation="h",
                    marker_color=C_RED,
                ))
                fig_bars.update_layout(
                    barmode="stack",
                    height=320,
                    margin=dict(t=10, b=0),
                    xaxis_title="USD",
                    legend=dict(orientation="h", yanchor="bottom", y=1.02),
                )
                st.plotly_chart(fig_bars, use_container_width=True)

        # Unfunded activities (zero partner contributions)
        st.divider()
        st.subheader("Unfunded Activities (Zero Partner Contributions)")
        unfunded = [g for g in gaps if g["total_support"] == 0 and g["total_cost"] > 0]
        if unfunded:
            uf_cols = st.columns(3)
            for i, act in enumerate(unfunded):
                with uf_cols[i % 3]:
                    pct_color = C_RED if act["priority"] == "High" else C_AMBER
                    st.markdown(
                        f"""
                        <div style='border:1px solid {pct_color};border-radius:6px;
                             padding:10px;margin-bottom:10px;'>
                          <div style='font-weight:700;font-size:.85rem;margin-bottom:4px;
                               color:{C_DARK};'>{act['activity_name'][:70]}</div>
                          <div style='font-size:.75rem;color:#555;'>
                            {act['pillar'][:30]}</div>
                          <div style='font-weight:600;color:{pct_color};margin-top:6px;'>
                            ${act['total_cost']:,.0f}  ·  Priority: {act['priority']}</div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )
                    if st.button("🔍 Find Partners", key=f"fp_{act['id']}"):
                        st.session_state.pc_sel_act = act["id"]
                        st.session_state["pc_view_radio"] = "📋 By Activity"
                        st.rerun()
        else:
            st.success("All activities with a budget have at least one partner contribution.")
