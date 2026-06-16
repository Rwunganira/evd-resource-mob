"""
WHO EVD Partner Resource Mobilization Dashboard — Visualizations
Run: streamlit run dashboard.py
Data management (add/edit): http://localhost:5000
"""
import io
import os
import requests
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ─────────────────────────────────────────────────────────────────────
FLASK_URL = os.getenv("FLASK_URL", "http://localhost:5000")
API_BASE = f"{FLASK_URL}/api"
DATA_MGMT_URL = FLASK_URL

WHO_BLUE   = "#0093D5"
WHO_DARK   = "#003865"
GREEN      = "#1E8449"
AMBER      = "#D97706"
RED        = "#C0392B"
LIGHT_GREY = "#E8E8E8"

st.set_page_config(
    page_title="WHO EVD Dashboard",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(f"""
<style>
  [data-testid="stSidebar"] {{ background-color: {WHO_DARK}; }}
  [data-testid="stSidebar"] * {{ color: white !important; }}
  .card {{
    background:white; border-left:5px solid {WHO_BLUE};
    border-radius:6px; padding:12px 16px;
    box-shadow:0 2px 6px rgba(0,0,0,.10); margin-bottom:8px;
  }}
  .card-green  {{ border-left-color:{GREEN}  !important; }}
  .card-red    {{ border-left-color:{RED}    !important; }}
  .card-amber  {{ border-left-color:{AMBER}  !important; }}
  .phase-banner {{
    padding:8px 20px; border-radius:6px; font-size:1rem;
    font-weight:700; text-align:center; margin-bottom:12px; color:white;
  }}
  .alert-box {{
    background:#fff3cd; border-left:4px solid #ffc107;
    padding:8px 14px; border-radius:4px; margin:4px 0;
  }}
  .offline-banner {{
    background:#fff0f0; border-left:4px solid {RED};
    padding:12px 16px; border-radius:4px; margin-bottom:16px;
    font-weight:600;
  }}
</style>
""", unsafe_allow_html=True)


# ── Session state ──────────────────────────────────────────────────────────────
if "last_refresh" not in st.session_state:
    st.session_state["last_refresh"] = datetime.now()


# ── API helpers ────────────────────────────────────────────────────────────────
@st.cache_data(ttl=60)
def api_get(endpoint: str, params: dict | None = None):
    try:
        r = requests.get(f"{API_BASE}/{endpoint.lstrip('/')}", params=params, timeout=5)
        r.raise_for_status()
        return r.json().get("data"), None
    except requests.exceptions.ConnectionError:
        return None, "Flask server is not running."
    except Exception as exc:
        return None, str(exc)


def offline_banner(msg: str):
    st.markdown(
        f"<div class='offline-banner'>⚠️ Backend unavailable — showing cached data or empty state.<br>"
        f"<small>{msg}</small><br><small>Start Flask: <code>python app.py</code></small></div>",
        unsafe_allow_html=True,
    )


def fmt_usd(val):
    if val >= 1_000_000:
        return f"${val/1_000_000:.1f}M"
    if val >= 1_000:
        return f"${val/1_000:.0f}K"
    return f"${val:,.0f}"


def coverage_color(pct):
    if pct >= 75:
        return GREEN
    if pct >= 50:
        return AMBER
    return RED


def refresh_button(key=""):
    if st.button("↻ Refresh", key=f"refresh_{key}"):
        st.cache_data.clear()
        st.session_state["last_refresh"] = datetime.now()
        st.rerun()
    elapsed = (datetime.now() - st.session_state["last_refresh"]).seconds // 60
    st.caption(f"Last refreshed: {elapsed} min ago")


# ── Excel export helper ────────────────────────────────────────────────────────
def make_excel(sheets: dict[str, tuple[list, list[list]]]) -> bytes:
    """sheets = {sheet_name: (headers, rows)}"""
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill(start_color="003865", end_color="003865", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet_name, (headers, rows) in sheets.items():
        ws = wb.create_sheet(sheet_name[:31])
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.row_dimensions[1].height = 20
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border
        for col in ws.columns:
            max_w = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    # Outbreak status banner
    summary_sb, _err_sb = api_get("dashboard/summary")
    phase_sb = (summary_sb or {}).get("current_phase") if summary_sb else None
    latest_sr_sb = (summary_sb or {}).get("latest_sitrep") if summary_sb else None

    if phase_sb:
        phase_name = phase_sb["phase_name"]
        phase_colors = {
            "Alert": RED, "Mobilization": AMBER,
            "Response": GREEN, "Scale-down": WHO_BLUE, "End": "#555",
        }
        bg = phase_colors.get(phase_name, WHO_BLUE)
        st.markdown(
            f"<div style='background:{bg};padding:8px;border-radius:4px;"
            f"text-align:center;margin-bottom:6px;'>"
            f"<span style='color:white;font-weight:800;font-size:.9rem;'>"
            f"● OUTBREAK {phase_name.upper()}</span></div>",
            unsafe_allow_html=True,
        )
    if latest_sr_sb:
        st.markdown(
            f"<div style='color:#a0c4e8;font-size:.72rem;padding-bottom:4px;'>"
            f"Last sitrep: {latest_sr_sb.get('report_date','—')}</div>",
            unsafe_allow_html=True,
        )

    st.markdown("""
    <div style='text-align:center;padding:6px 0 4px;'>
      <div style='font-size:1.8rem;'>🏥</div>
      <div style='font-size:1.05rem;font-weight:800;'>WHO EOC</div>
      <div style='font-size:.72rem;color:#a0c4e8;'>EVD Response — DRC</div>
    </div>
    <hr style='border-color:#1a5276;margin:6px 0;'>
    """, unsafe_allow_html=True)

    # Quick stats
    if summary_sb:
        p = summary_sb.get("partners", {})
        f = summary_sb.get("funding", {})
        day = summary_sb.get("outbreak_day", 0)
        st.markdown(
            f"<div style='background:#1a4a6b;border-radius:4px;padding:8px;margin:6px 0;"
            f"font-size:.75rem;color:white;text-align:center;'>"
            f"👥 {p.get('total',0)} partners &nbsp;|&nbsp; "
            f"💰 {fmt_usd(f.get('total_committed_usd',0))} &nbsp;|&nbsp; "
            f"📅 Day {day}</div>",
            unsafe_allow_html=True,
        )

    page = st.radio(
        "Navigation",
        [
            "Overview / Command Center",
            "Partner Analysis",
            "Funding Gap Tracker",
            "Partner Compliance",
            "Resource Mobilization",
            "4W Matrix",
            "Epidemic Curve",
            "Export & Import",
        ],
        label_visibility="collapsed",
    )

    st.markdown("<hr style='border-color:#1a5276;'>", unsafe_allow_html=True)
    if st.button("↻ Refresh All", key="sidebar_refresh"):
        st.cache_data.clear()
        st.session_state["last_refresh"] = datetime.now()
        st.rerun()
    st.markdown(
        f"[Open Data Management ↗]({DATA_MGMT_URL})",
        unsafe_allow_html=False,
    )


# ══════════════════════════════════════════════════════════════════════════════
# PAGE 1 — OVERVIEW / COMMAND CENTER
# ══════════════════════════════════════════════════════════════════════════════
if page == "Overview / Command Center":
    st.title("EVD Response — Command Center")
    col_r, _ = st.columns([1, 6])
    with col_r:
        refresh_button("overview")

    with st.spinner("Loading dashboard summary…"):
        summary, err = api_get("dashboard/summary")
    if err:
        offline_banner(err)
        st.stop()

    phase = summary.get("current_phase")
    if phase:
        phase_name = phase["phase_name"]
        phase_colors = {
            "Alert": RED, "Mobilization": AMBER,
            "Response": GREEN, "Scale-down": WHO_BLUE, "End": "#555",
        }
        bg = phase_colors.get(phase_name, WHO_BLUE)
        st.markdown(
            f"<div class='phase-banner' style='background:{bg};'>"
            f"OUTBREAK PHASE: {phase_name.upper()}&nbsp;&nbsp;|&nbsp;&nbsp;"
            f"Day {summary['outbreak_day']} of EVD Response</div>",
            unsafe_allow_html=True,
        )

    # Key metrics row
    col1, col2, col3, col4, col5 = st.columns(5)
    f = summary.get("funding", {})
    p = summary.get("partners", {})
    sr = summary.get("latest_sitrep") or {}
    col1.metric("Partners Mobilized", p.get("total", 0), delta=f"+{p.get('active', 0)} active")
    col2.metric("Funding Committed", fmt_usd(f.get("total_committed_usd", 0)))
    col3.metric("Funding Gap", fmt_usd(f.get("gap_usd", 0)),
                delta=f"{f.get('coverage_percent', 0):.0f}% covered", delta_color="off")
    col4.metric("ETUs Operational", summary.get("etus_operational", 0))
    col5.metric("Outbreak Day", summary.get("outbreak_day", 0))

    st.divider()
    col_left, col_right = st.columns([1.4, 1])

    with col_left:
        st.subheader("Latest Situation Report")
        if sr:
            c1, c2, c3 = st.columns(3)
            c1.metric("Confirmed Cases", sr.get("confirmed_cases", 0))
            c2.metric("Deaths", sr.get("deaths", 0))
            c3.metric("CFR %", f"{sr.get('cfr_percent', 0)}%")
            st.caption(
                f"Report date: {sr.get('report_date')}  |  "
                f"HCW affected: {sr.get('healthcare_workers_affected', 0)}"
            )
            if sr.get("notes"):
                st.info(sr["notes"])
        else:
            st.warning("No situation reports available.")

        # Funding coverage gauge
        st.subheader("Funding Coverage (Resource Commitments)")
        coverage = f.get("coverage_percent", 0)
        fig_gauge = go.Figure(go.Indicator(
            mode="gauge+number+delta",
            value=coverage,
            number={"suffix": "%"},
            delta={"reference": 100, "increasing": {"color": GREEN}},
            gauge={
                "axis": {"range": [0, 100]},
                "bar": {"color": coverage_color(coverage)},
                "steps": [
                    {"range": [0, 50],  "color": "#ffd0d0"},
                    {"range": [50, 75], "color": "#fff0c0"},
                    {"range": [75, 100],"color": "#d0f0d0"},
                ],
                "threshold": {"line": {"color": RED, "width": 3}, "value": 80},
            },
            title={"text": "Funding Coverage"},
        ))
        fig_gauge.update_layout(height=220, margin=dict(t=40, b=0, l=20, r=20))
        st.plotly_chart(fig_gauge, use_container_width=True)

    with col_right:
        # Partner compliance badge
        st.subheader("Partner Reporting This Week")
        reporting = summary.get("partners_reporting_this_week", 0)
        overdue = summary.get("partners_overdue_this_week", 0)
        total_active = p.get("total", 1)
        comp_pct = round(reporting / total_active * 100, 0) if total_active else 0
        badge_color = GREEN if comp_pct >= 80 else AMBER if comp_pct >= 50 else RED
        st.markdown(
            f"<div style='background:{badge_color};color:white;border-radius:8px;"
            f"padding:16px;text-align:center;font-weight:700;font-size:1.2rem;'>"
            f"{reporting} of {total_active} partners reported<br>"
            f"<span style='font-size:.85rem;font-weight:400;'>{comp_pct:.0f}% compliance"
            f"{'  ✓' if comp_pct >= 80 else '  ⚠' if comp_pct >= 50 else '  ✗'}</span>"
            f"</div>",
            unsafe_allow_html=True,
        )
        if overdue:
            st.caption(f"⚠️ {overdue} partners have not yet submitted this week")

        # Most active location
        most_active = summary.get("most_active_location")
        if most_active:
            st.markdown(
                f"<div class='card' style='margin-top:12px;'>"
                f"<div style='font-size:.75rem;color:#666;text-transform:uppercase;letter-spacing:.05em;'>Most Active Location</div>"
                f"<div style='font-size:1.4rem;font-weight:700;color:{WHO_DARK};'>{most_active}</div>"
                f"</div>",
                unsafe_allow_html=True,
            )

        # Alerts
        st.subheader("Reporting Alerts")
        due_data, _ = api_get("resources/reporting-due", {"days": 3})
        if due_data:
            overdue_r = due_data.get("overdue", [])
            due_soon = due_data.get("due_soon", [])
            if overdue_r:
                for r in overdue_r:
                    st.markdown(
                        f"<div class='alert-box'>⚠️ <b>{r['partner_name']}</b> — "
                        f"{r['resource_type']} overdue since {r['next_report_due']}</div>",
                        unsafe_allow_html=True,
                    )
            if due_soon:
                for r in due_soon:
                    st.markdown(
                        f"<div class='alert-box' style='background:#e8f4fd;border-color:{WHO_BLUE};'>"
                        f"📋 <b>{r['partner_name']}</b> — due {r['next_report_due']}</div>",
                        unsafe_allow_html=True,
                    )
            if not overdue_r and not due_soon:
                st.success("No overdue reports in the next 3 days.")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE 2 — PARTNER ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Partner Analysis":
    st.title("Partner Analysis")
    col_r, _ = st.columns([1, 6])
    with col_r:
        refresh_button("partners")

    with st.spinner("Loading partners…"):
        partners, err = api_get("partners")
    if err:
        offline_banner(err)
        st.stop()

    st.info(f"To add or edit partners, use the [Data Management app]({DATA_MGMT_URL}/partners).")
    df = pd.DataFrame(partners)

    col1, col2, col3 = st.columns(3)
    with col1:
        sel_type = st.selectbox("Filter by Type", ["All"] + sorted(df["partner_type"].unique()))
    with col2:
        sel_status = st.selectbox("Filter by Status", ["All"] + sorted(df["status"].unique()))
    with col3:
        sel_country = st.selectbox("Filter by Country", ["All"] + sorted(df["country"].unique()))

    filtered = df.copy()
    if sel_type != "All":
        filtered = filtered[filtered["partner_type"] == sel_type]
    if sel_status != "All":
        filtered = filtered[filtered["status"] == sel_status]
    if sel_country != "All":
        filtered = filtered[filtered["country"] == sel_country]

    st.markdown(f"**{len(filtered)} partners** matching filters")
    st.dataframe(
        filtered[["name", "partner_type", "country", "status", "contact_person", "resource_count", "activity_count"]],
        use_container_width=True, hide_index=True,
        column_config={
            "name": "Partner", "partner_type": "Type", "country": "Country",
            "status": "Status", "contact_person": "Contact",
            "resource_count": "Resources", "activity_count": "Activities",
        },
    )

    col_a, col_b = st.columns(2)
    with col_a:
        fig_pie = px.pie(
            df, names="status", title="Partner Status Distribution",
            color_discrete_map={
                "Active": GREEN, "Negotiating": WHO_BLUE,
                "Pipeline": AMBER, "Inactive": "#aaa",
            },
            hole=0.4,
        )
        fig_pie.update_layout(height=320, margin=dict(t=40, b=0))
        st.plotly_chart(fig_pie, use_container_width=True)
    with col_b:
        type_counts = df["partner_type"].value_counts().reset_index()
        type_counts.columns = ["Type", "Count"]
        fig_bar = px.bar(type_counts, x="Count", y="Type", orientation="h",
                         color_discrete_sequence=[WHO_BLUE], title="By Partner Type")
        fig_bar.update_layout(height=320, margin=dict(t=40, b=0))
        st.plotly_chart(fig_bar, use_container_width=True)

    st.divider()
    st.subheader("Partner Detail")
    partner_names = {p["name"]: p["id"] for p in partners}
    selected_name = st.selectbox("Select partner", ["— select —"] + list(partner_names))
    if selected_name != "— select —":
        detail, _ = api_get(f"partners/{partner_names[selected_name]}")
        if detail:
            c1, c2 = st.columns(2)
            with c1:
                st.markdown(f"**Type:** {detail['partner_type']}  |  **Country:** {detail['country']}")
                st.markdown(f"**Status:** {detail['status']}  |  **Contact:** {detail.get('contact_person','')} ({detail.get('contact_email','')})")
                st.markdown("**Resources:**")
                if detail["resources"]:
                    st.dataframe(pd.DataFrame(detail["resources"])[["resource_type","amount","currency","status"]],
                                 hide_index=True, use_container_width=True)
                else:
                    st.caption("No resources logged.")
            with c2:
                st.markdown("**Activities:**")
                if detail["activities"]:
                    st.dataframe(pd.DataFrame(detail["activities"])[["activity_type","location","status"]],
                                 hide_index=True, use_container_width=True)
                else:
                    st.caption("No activities logged.")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE 3 — FUNDING GAP TRACKER
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Funding Gap Tracker":
    st.title("Funding Gap Tracker")
    col_r, _ = st.columns([1, 6])
    with col_r:
        refresh_button("funding")

    with st.spinner("Loading funding data…"):
        gap_summary, err = api_get("funding-gap/summary")
    if err:
        offline_banner(err)
        st.stop()

    if not gap_summary:
        st.warning("No funding gap data available.")
        st.stop()

    pillars = gap_summary.get("pillars", [])
    total_req = gap_summary["total_required_usd"]
    total_funded = gap_summary["total_funded_usd"]
    total_gap = gap_summary["total_gap_usd"]
    coverage = gap_summary["coverage_pct"]

    col_left, col_right = st.columns(2)

    with col_left:
        st.subheader("Overall Response Funding")
        fig_g = go.Figure(go.Indicator(
            mode="gauge+number+delta",
            value=coverage,
            number={"suffix": "%"},
            delta={"reference": 100},
            gauge={
                "axis": {"range": [0, 100]},
                "bar": {"color": coverage_color(coverage)},
                "steps": [
                    {"range": [0, 50],  "color": "#ffd0d0"},
                    {"range": [50, 75], "color": "#fff0c0"},
                    {"range": [75, 100],"color": "#d0f0d0"},
                ],
                "threshold": {"line": {"color": RED, "width": 3}, "value": 75},
            },
            title={"text": "Overall Coverage"},
        ))
        fig_g.update_layout(height=260, margin=dict(t=50, b=0))
        st.plotly_chart(fig_g, use_container_width=True)

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Required", fmt_usd(total_req))
        m2.metric("Funded", fmt_usd(total_funded))
        m3.metric("Gap", fmt_usd(total_gap))
        m4.metric("Coverage", f"{coverage}%")

    with col_right:
        st.subheader("Funded vs Required by Pillar")
        df_p = pd.DataFrame(pillars).sort_values("gap_usd", ascending=False)

        fig_h = go.Figure()
        fig_h.add_trace(go.Bar(
            y=df_p["pillar_name"],
            x=df_p["amount_required_usd"],
            name="Required",
            orientation="h",
            marker_color=LIGHT_GREY,
        ))
        fig_h.add_trace(go.Bar(
            y=df_p["pillar_name"],
            x=df_p["amount_funded_usd"],
            name="Funded",
            orientation="h",
            marker_color=WHO_BLUE,
        ))
        fig_h.update_layout(
            barmode="overlay",
            height=320,
            margin=dict(t=10, b=0),
            legend=dict(orientation="h", yanchor="bottom", y=1.02),
            xaxis_title="USD",
        )
        st.plotly_chart(fig_h, use_container_width=True)

    # Pillar details table
    st.subheader("Pillar Details")

    def _status_label(pct):
        if pct >= 75:
            return "Adequate"
        if pct >= 50:
            return "At Risk"
        return "Critical"

    def _status_color(val):
        if val == "Adequate":
            return f"background-color: #d4edda; color: #155724;"
        if val == "At Risk":
            return f"background-color: #fff3cd; color: #856404;"
        return f"background-color: #f8d7da; color: #721c24;"

    table_df = pd.DataFrame([{
        "Pillar": g["pillar_name"],
        "Required (USD)": g["amount_required_usd"],
        "Funded (USD)": g["amount_funded_usd"],
        "Gap (USD)": g["gap_usd"],
        "Coverage %": g["coverage_pct"],
        "Status": _status_label(g["coverage_pct"]),
    } for g in pillars])

    styled = table_df.style.applymap(_status_color, subset=["Status"]).format({
        "Required (USD)": "${:,.0f}",
        "Funded (USD)":   "${:,.0f}",
        "Gap (USD)":      "${:,.0f}",
        "Coverage %":     "{:.1f}%",
    })
    st.dataframe(styled, use_container_width=True, hide_index=True)



# ══════════════════════════════════════════════════════════════════════════════
# PAGE 4 — PARTNER COMPLIANCE TRACKER
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Partner Compliance":
    st.title("Partner Compliance Tracker")
    col_r, _ = st.columns([1, 6])
    with col_r:
        refresh_button("compliance")

    with st.spinner("Loading compliance data…"):
        comp_data, err = api_get("partner-reports/compliance")
    if err:
        offline_banner(err)
        st.stop()

    if not comp_data:
        st.warning("No compliance data available.")
        st.stop()

    this_submitted = comp_data["this_week_submitted"]
    this_overdue = comp_data["this_week_overdue"]
    total_active = comp_data["total_active_partners"]
    comp_pct = round(this_submitted / total_active * 100, 0) if total_active else 0

    # Summary banner
    banner_color = GREEN if comp_pct >= 80 else AMBER if comp_pct >= 50 else RED
    st.markdown(
        f"<div style='background:{banner_color};color:white;border-radius:8px;"
        f"padding:12px 20px;font-weight:700;font-size:1rem;margin-bottom:16px;'>"
        f"✅ {this_submitted} compliant &nbsp;|&nbsp; "
        f"⚠️ {this_overdue} overdue &nbsp;|&nbsp; "
        f"📊 {comp_pct:.0f}% compliance rate this week</div>",
        unsafe_allow_html=True,
    )

    # Partner status cards
    st.subheader("This Week's Status")
    by_partner = comp_data["by_partner"]
    weeks = comp_data["weeks"]
    current_week = weeks[-1] if weeks else None

    cols = st.columns(4)
    for i, p_data in enumerate(by_partner):
        submitted_this_week = p_data["by_week"].get(current_week, False) if current_week else False
        card_color = GREEN if submitted_this_week else RED
        icon = "✅" if submitted_this_week else "⚠️"
        label = "Submitted" if submitted_this_week else "Overdue"
        with cols[i % 4]:
            st.markdown(
                f"<div class='card' style='border-left-color:{card_color};'>"
                f"<div style='font-weight:700;font-size:.9rem;'>{p_data['partner_name']}</div>"
                f"<div style='color:{card_color};font-size:.85rem;'>{icon} {label}</div>"
                f"<div style='font-size:.75rem;color:#666;'>{p_data['compliance_pct']:.0f}% (4-wk)</div>"
                f"</div>",
                unsafe_allow_html=True,
            )

    # Historical heatmap
    st.divider()
    st.subheader("Historical Compliance (Last 4 Weeks)")
    heat_rows = []
    for p_data in by_partner:
        row = {"Partner": p_data["partner_name"]}
        for w in weeks:
            row[w[:10]] = "✅" if p_data["by_week"].get(w, False) else "❌"
        row["4-Wk %"] = f"{p_data['compliance_pct']:.0f}%"
        heat_rows.append(row)
    st.dataframe(pd.DataFrame(heat_rows), use_container_width=True, hide_index=True)

    # Compliance trend
    st.subheader("Compliance Trend")
    by_week = comp_data["by_week"]
    trend_df = pd.DataFrame(by_week)
    if not trend_df.empty:
        fig_trend = px.line(
            trend_df, x="week", y="compliance_pct",
            markers=True,
            labels={"week": "Week", "compliance_pct": "Compliance %"},
            color_discrete_sequence=[WHO_BLUE],
        )
        fig_trend.add_hline(y=80, line_dash="dot", line_color=GREEN,
                            annotation_text="80% target")
        fig_trend.update_layout(height=260, margin=dict(t=10, b=0),
                                yaxis=dict(range=[0, 110]))
        st.plotly_chart(fig_trend, use_container_width=True)

    # Reminder section
    st.divider()
    st.subheader("Send Reminder to Overdue Partners")
    if st.button("Show Overdue Partner Contacts"):
        overdue_data, _ = api_get("partner-reports/overdue")
        if overdue_data:
            overdue_partners = overdue_data.get("overdue_partners", [])
            if overdue_partners:
                for op in overdue_partners:
                    st.markdown(
                        f"**{op['name']}** — {op.get('contact_person','—')} "
                        f"— [{op.get('contact_email','—')}](mailto:{op.get('contact_email','')})"
                    )
            else:
                st.success("No overdue partners this week.")


# ══════════════════════════════════════════════════════════════════════════════
# PAGE 5 — RESOURCE MOBILIZATION
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Resource Mobilization":
    st.title("Resource Mobilization")
    col_r, _ = st.columns([1, 6])
    with col_r:
        refresh_button("resources")

    with st.spinner("Loading resources…"):
        resources, err = api_get("resources")
    if err:
        offline_banner(err)
        st.stop()

    st.info(f"To add or edit resources, use the [Data Management app]({DATA_MGMT_URL}/resources).")
    df = pd.DataFrame(resources)
    gap_data, _ = api_get("resources/funding-gap")
    by_type_data, _ = api_get("resources/total-by-type")

    if gap_data:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total Committed (USD)", fmt_usd(gap_data["total_committed_usd"]))
        c2.metric("Total Deployed (USD)", fmt_usd(gap_data["total_deployed_usd"]))
        c3.metric("Funding Gap", fmt_usd(gap_data["gap_usd"]))
        c4.metric("Coverage", f"{gap_data['coverage_percent']}%")

    col1, col2 = st.columns(2)
    with col1:
        if by_type_data:
            st.subheader("Resources by Type (USD)")
            type_df = pd.DataFrame(list(by_type_data.items()), columns=["Type", "Amount (USD)"])
            fig_h = px.bar(type_df.sort_values("Amount (USD)"),
                           x="Amount (USD)", y="Type", orientation="h",
                           color_discrete_sequence=[WHO_BLUE])
            fig_h.update_layout(height=280, margin=dict(t=10, b=0))
            st.plotly_chart(fig_h, use_container_width=True)

    with col2:
        if gap_data:
            committed = gap_data["total_committed_usd"]
            needed = gap_data["estimated_need_usd"]
            fig_g = go.Figure(go.Indicator(
                mode="gauge+number",
                value=committed / 1_000_000,
                number={"suffix": "M USD"},
                gauge={
                    "axis": {"range": [0, needed / 1_000_000]},
                    "bar": {"color": WHO_BLUE},
                    "threshold": {"line": {"color": RED, "width": 3}, "value": needed / 1_000_000},
                },
                title={"text": f"Funding vs Need ({fmt_usd(needed)})"},
            ))
            fig_g.update_layout(height=260, margin=dict(t=50, b=0))
            st.plotly_chart(fig_g, use_container_width=True)

    if not df.empty and "commitment_date" in df.columns:
        st.subheader("Funding Commitments Over Time")
        funding_df = df[df["resource_type"] == "Funding"].copy()
        if not funding_df.empty:
            funding_df["commitment_date"] = pd.to_datetime(funding_df["commitment_date"])
            funding_df = funding_df.sort_values("commitment_date")
            funding_df["cumulative"] = funding_df["amount"].cumsum()
            fig_line = px.line(funding_df, x="commitment_date", y="cumulative",
                               markers=True, color_discrete_sequence=[WHO_BLUE])
            fig_line.update_layout(height=240, margin=dict(t=10, b=0))
            st.plotly_chart(fig_line, use_container_width=True)

    st.subheader("All Resource Commitments")
    STATUS_COLORS = {
        "Deployed":    "background-color:#d4edda",
        "Committed":   "background-color:#fff3cd",
        "Pipeline":    "background-color:#ffe0b2",
        "Under Review":"background-color:#f8d7da",
    }
    display_cols = ["partner_name", "resource_type", "description", "amount", "currency", "status", "next_report_due"]
    display_df = df[[c for c in display_cols if c in df.columns]]
    st.dataframe(
        display_df.style.map(lambda v: STATUS_COLORS.get(v, ""), subset=["status"])
        if "status" in display_df.columns else display_df,
        use_container_width=True, hide_index=True,
    )


# ══════════════════════════════════════════════════════════════════════════════
# PAGE 6 — 4W MATRIX
# ══════════════════════════════════════════════════════════════════════════════
elif page == "4W Matrix":
    st.title("4W Matrix — Who does What Where When")
    col_r, _ = st.columns([1, 6])
    with col_r:
        refresh_button("4w")

    with st.spinner("Loading 4W data…"):
        matrix_data, err = api_get("activities/4w-matrix")
    if err:
        offline_banner(err)
        st.stop()

    if not matrix_data:
        st.warning("No activity data available.")
        st.info(f"Add activities via the [Data Management app]({DATA_MGMT_URL}/activities).")
        st.stop()

    flat = matrix_data.get("flat", [])
    act_df = pd.DataFrame(flat) if flat else pd.DataFrame()

    # Filters
    with st.expander("Filters", expanded=True):
        fc1, fc2, fc3, fc4 = st.columns(4)
        all_districts = sorted(set(r.get("district") or "" for r in flat if r.get("district")))
        all_types = sorted(set(r.get("activity_type", "") for r in flat))
        all_statuses = ["Planned", "Ongoing", "Completed"]
        f_district = fc1.selectbox("District", ["All"] + all_districts)
        f_type     = fc2.selectbox("Activity Type", ["All"] + all_types)
        f_status   = fc3.selectbox("Status", ["All"] + all_statuses)
        if not act_df.empty and "partner_name" in act_df.columns:
            all_partners = sorted(act_df["partner_name"].dropna().unique())
            f_partner  = fc4.selectbox("Partner", ["All"] + all_partners)
        else:
            f_partner = "All"

    filtered_df = act_df.copy()
    if not filtered_df.empty:
        if f_district != "All":
            filtered_df = filtered_df[filtered_df["district"] == f_district]
        if f_type != "All":
            filtered_df = filtered_df[filtered_df["activity_type"] == f_type]
        if f_status != "All":
            filtered_df = filtered_df[filtered_df["status"] == f_status]
        if f_partner != "All":
            filtered_df = filtered_df[filtered_df["partner_name"] == f_partner]

    # 3W summary table
    st.subheader("Activity Matrix")
    matrix = matrix_data.get("matrix", {})
    locations = matrix_data.get("locations", [])
    activity_types = matrix_data.get("activity_types", [])
    rows_3w = []
    for loc in locations:
        row = {"Location": loc}
        for atype in activity_types:
            partners_list = matrix.get(loc, {}).get(atype, [])
            row[atype] = ", ".join(partners_list) if partners_list else ""
        rows_3w.append(row)
    st.dataframe(pd.DataFrame(rows_3w).set_index("Location"), use_container_width=True)

    # Map or bar chart
    PROVINCE_COORDS = {
        "Equateur Province": {"lat": 0.5, "lon": 22.0},
        "Nord-Ubangi":       {"lat": 3.5, "lon": 21.5},
        "Sud-Ubangi":        {"lat": 2.5, "lon": 21.0},
        "Mbandaka":          {"lat": 0.048, "lon": 18.26},
        "Bikoro":            {"lat": -0.7, "lon": 18.1},
        "Ingende":           {"lat": -0.26, "lon": 18.94},
        "Bolomba":           {"lat": -0.4, "lon": 19.37},
    }
    if not filtered_df.empty:
        filtered_df["lat"] = filtered_df["location"].map(lambda x: PROVINCE_COORDS.get(x, {}).get("lat"))
        filtered_df["lon"] = filtered_df["location"].map(lambda x: PROVINCE_COORDS.get(x, {}).get("lon"))
        map_df = filtered_df.dropna(subset=["lat", "lon"])

        col1, col2 = st.columns([2, 1])
        with col1:
            st.subheader("Activity Locations")
            if not map_df.empty:
                sz_col = "beneficiaries_reached"
                map_df = map_df.copy()
                map_df[sz_col] = map_df[sz_col].fillna(0).clip(lower=1)
                fig_map = px.scatter_mapbox(
                    map_df, lat="lat", lon="lon",
                    color="activity_type", size=sz_col,
                    size_max=30,
                    hover_name="partner_name",
                    hover_data={"location": True, "status": True,
                                "beneficiaries_reached": True, "lat": False, "lon": False},
                    zoom=4, height=400, mapbox_style="open-street-map",
                )
                fig_map.update_layout(margin=dict(t=0, b=0))
                st.plotly_chart(fig_map, use_container_width=True)
            else:
                st.info("No coordinate matches — showing location frequency instead.")
                loc_cnt = filtered_df["location"].value_counts().reset_index()
                loc_cnt.columns = ["Location", "Activities"]
                st.bar_chart(loc_cnt.set_index("Location"))

        with col2:
            st.subheader("By Type")
            tc = filtered_df["activity_type"].value_counts().reset_index()
            tc.columns = ["Type", "Count"]
            fig_tc = px.bar(tc, x="Count", y="Type", orientation="h",
                            color_discrete_sequence=[WHO_BLUE])
            fig_tc.update_layout(height=200, margin=dict(t=10, b=0))
            st.plotly_chart(fig_tc, use_container_width=True)

            st.subheader("By Status")
            sc = filtered_df["status"].value_counts().reset_index()
            sc.columns = ["Status", "Count"]
            fig_sc = px.pie(sc, names="Status", values="Count", hole=0.4,
                            color_discrete_map={"Ongoing": GREEN, "Planned": WHO_BLUE, "Completed": "#888"})
            fig_sc.update_layout(height=200, margin=dict(t=10, b=0))
            st.plotly_chart(fig_sc, use_container_width=True)

    # Export
    st.divider()
    if not filtered_df.empty:
        export_data, _ = api_get("activities/4w-matrix/export")
        if export_data:
            headers = ["Partner", "Activity Type", "Location", "District",
                       "Start Date", "End Date", "Status", "Beneficiaries Reached", "Description"]
            rows_xl = [
                [r["partner_name"], r["activity_type"], r["location"], r.get("district",""),
                 r.get("start_date",""), r.get("end_date",""), r["status"],
                 r.get("beneficiaries_reached",0), r.get("description","")]
                for r in export_data
            ]
            xl_bytes = make_excel({"4W Matrix": (headers, rows_xl)})
            st.download_button(
                "Download 4W Matrix (Excel)", xl_bytes,
                file_name=f"4W_Matrix_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )


# ══════════════════════════════════════════════════════════════════════════════
# PAGE 7 — EPIDEMIC CURVE
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Epidemic Curve":
    st.title("Epidemic Curve & Situation Reports")
    col_r, _ = st.columns([1, 6])
    with col_r:
        refresh_button("epi")

    with st.spinner("Loading situation reports…"):
        sitreps, err = api_get("sitreps")
    if err:
        offline_banner(err)
        st.stop()

    st.info(f"To submit a new situation report, use the [Data Management app]({DATA_MGMT_URL}/sitreps).")

    if not sitreps:
        st.warning("No situation reports available yet.")
        st.stop()

    df_sr = pd.DataFrame(sitreps)
    df_sr["report_date"] = pd.to_datetime(df_sr["report_date"])
    df_sr = df_sr.sort_values("report_date")

    # Epidemic curve with CFR on second Y-axis
    st.subheader("Epidemic Curve with Case Fatality Rate")
    fig_epi = go.Figure()
    fig_epi.add_trace(go.Scatter(
        x=df_sr["report_date"], y=df_sr["confirmed_cases"],
        name="Confirmed Cases", mode="lines+markers",
        line=dict(color=AMBER, width=2),
    ))
    fig_epi.add_trace(go.Scatter(
        x=df_sr["report_date"], y=df_sr["deaths"],
        name="Deaths", mode="lines+markers",
        line=dict(color=RED, width=2),
    ))
    fig_epi.add_trace(go.Scatter(
        x=df_sr["report_date"], y=df_sr["cfr_percent"],
        name="CFR %", mode="lines+markers",
        line=dict(color=WHO_BLUE, width=2, dash="dot"),
        yaxis="y2",
    ))

    # Annotation markers from activities (key events)
    activities_ann, _ = api_get("activities")
    if activities_ann:
        key_events = [
            a for a in activities_ann
            if a.get("activity_type") in ("ETU Operations", "Vaccination")
            and a.get("start_date")
        ]
        for event in key_events[:3]:  # limit annotations
            fig_epi.add_vline(
                x=event["start_date"],
                line_dash="dot", line_color="#888", line_width=1,
                annotation_text=f"{event['activity_type'][:8]}",
                annotation_position="top",
            )

    fig_epi.update_layout(
        height=360, margin=dict(t=10, b=0),
        xaxis_title="Date",
        yaxis=dict(title="Cases / Deaths"),
        yaxis2=dict(title="CFR %", overlaying="y", side="right",
                    range=[0, max(df_sr["cfr_percent"].max() * 1.2, 10)]),
        legend=dict(orientation="h", yanchor="bottom", y=1.02),
    )
    st.plotly_chart(fig_epi, use_container_width=True)

    # CFR and funding charts
    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Case Fatality Rate Trend")
        fig_cfr = px.line(df_sr, x="report_date", y="cfr_percent",
                          markers=True, color_discrete_sequence=[RED])
        fig_cfr.update_layout(height=240, margin=dict(t=10, b=0))
        st.plotly_chart(fig_cfr, use_container_width=True)
    with col2:
        st.subheader("Funding Mobilized (USD)")
        fig_fund = px.bar(df_sr, x="report_date", y="total_funding_mobilized",
                          color_discrete_sequence=[WHO_BLUE])
        fig_fund.update_layout(height=240, margin=dict(t=10, b=0))
        st.plotly_chart(fig_fund, use_container_width=True)

    # Latest sitrep metrics
    latest = sitreps[0]
    st.subheader(f"Latest SitRep — Day {latest['outbreak_day']} ({latest['report_date']})")
    lc1, lc2, lc3, lc4, lc5 = st.columns(5)
    lc1.metric("Confirmed Cases", latest["confirmed_cases"])
    lc2.metric("Deaths", latest["deaths"])
    lc3.metric("CFR", f"{latest['cfr_percent']}%")
    lc4.metric("HCW Affected", latest["healthcare_workers_affected"])
    lc5.metric("ETUs Operational", latest["etus_operational"])
    if latest.get("notes"):
        st.info(latest["notes"])

    # SitRep comparison table — last 3
    if len(sitreps) >= 2:
        st.divider()
        st.subheader("SitRep Comparison (Last 3 Reports)")
        n0 = sitreps[0]
        n1 = sitreps[1] if len(sitreps) > 1 else None
        n2 = sitreps[2] if len(sitreps) > 2 else None

        indicators = [
            ("confirmed_cases",          "Confirmed Cases",        True),
            ("deaths",                   "Deaths",                 True),
            ("cfr_percent",              "CFR %",                  True),
            ("healthcare_workers_affected","HCW Affected",         True),
            ("etus_operational",         "ETUs Operational",       False),
            ("total_funding_mobilized",  "Funding Mobilized (USD)",False),
        ]
        comp_rows = []
        for field, label, bad_if_up in indicators:
            v0 = n0.get(field, 0)
            v1 = n1.get(field, 0) if n1 else None
            v2 = n2.get(field, 0) if n2 else None
            if v1 is not None and v0 != v1:
                arrow = "↑" if v0 > v1 else "↓"
                change = f"{arrow} {abs(v0 - v1):,.1f}"
            else:
                change = "—"
            comp_rows.append({
                "Indicator": label,
                f"N-2 ({n2['report_date'] if n2 else '—'})": f"{v2:,.1f}" if v2 is not None else "—",
                f"N-1 ({n1['report_date'] if n1 else '—'})": f"{v1:,.1f}" if v1 is not None else "—",
                f"Latest ({n0['report_date']})": f"{v0:,.1f}",
                "Change": change,
            })
        st.dataframe(pd.DataFrame(comp_rows), use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# PAGE 8 — EXPORT & IMPORT
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Export & Import":
    st.title("Export & Import")
    col_r, _ = st.columns([1, 6])
    with col_r:
        refresh_button("export")

    with st.spinner("Loading data…"):
        partners, _  = api_get("partners")
        resources, _ = api_get("resources")
        activities_raw, _ = api_get("activities")
        sitreps, _   = api_get("sitreps")
        pillars, _   = api_get("funding-gap")
        due_data, _  = api_get("resources/reporting-due", {"days": 14})

    tab_export, = st.tabs(["📥 Export"])

    # ── Export tab ─────────────────────────────────────────────────────────────
    with tab_export:
        # Reporting deadlines
        st.subheader("Upcoming Reporting Deadlines (14 days)")
        if due_data:
            overdue_r = due_data.get("overdue", [])
            due_soon = due_data.get("due_soon", [])
            if overdue_r:
                st.error(f"{len(overdue_r)} overdue reports")
                st.dataframe(
                    pd.DataFrame(overdue_r)[["partner_name","resource_type","next_report_due","reporting_frequency"]],
                    hide_index=True, use_container_width=True,
                )
            if due_soon:
                st.warning(f"{len(due_soon)} reports due within 14 days")
                st.dataframe(
                    pd.DataFrame(due_soon)[["partner_name","resource_type","next_report_due","reporting_frequency"]],
                    hide_index=True, use_container_width=True,
                )

        if resources:
            st.subheader("Resource Utilization Summary")
            r_df = pd.DataFrame(resources)
            st.dataframe(
                r_df.groupby(["resource_type","status"])["amount"].sum().reset_index(),
                hide_index=True, use_container_width=True,
            )

        st.divider()
        st.subheader("Download Exports")

        col1, col2, col3 = st.columns(3)

        # Partner Directory Excel
        with col1:
            if partners:
                headers = ["Name","Type","Country","Contact Person","Contact Email","Status","Resources","Activities"]
                rows_xl = [
                    [p["name"], p["partner_type"], p["country"],
                     p.get("contact_person",""), p.get("contact_email",""),
                     p["status"], p.get("resource_count",0), p.get("activity_count",0)]
                    for p in partners
                ]
                st.download_button(
                    "📋 Partner Directory (Excel)",
                    make_excel({"Partners": (headers, rows_xl)}),
                    file_name=f"Partner_Directory_{date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

        # Resource Commitments Excel
        with col2:
            if resources:
                headers = ["Partner","Type","Description","Amount","Currency",
                           "Status","Commitment Date","Deployment Date","Next Report Due"]
                rows_xl = [
                    [r["partner_name"], r["resource_type"], r.get("description",""),
                     r["amount"], r["currency"], r["status"],
                     r.get("commitment_date",""), r.get("deployment_date",""),
                     r.get("next_report_due","")]
                    for r in resources
                ]
                st.download_button(
                    "💰 Resource Commitments (Excel)",
                    make_excel({"Resources": (headers, rows_xl)}),
                    file_name=f"Resource_Commitments_{date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

        # Funding Gap Report Excel
        with col3:
            if pillars:
                headers = ["Pillar","Required (USD)","Funded (USD)","Gap (USD)","Coverage %","Notes"]
                rows_xl = [
                    [p["pillar_name"], p["amount_required_usd"], p["amount_funded_usd"],
                     p["gap_usd"], f"{p['coverage_pct']}%", p.get("notes","")]
                    for p in pillars
                ]
                st.download_button(
                    "📊 Funding Gap Report (Excel)",
                    make_excel({"Funding Gap": (headers, rows_xl)}),
                    file_name=f"Funding_Gap_Report_{date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

        col4, col5 = st.columns(3)[:2]

        # 4W Matrix Excel
        with col4:
            if activities_raw:
                headers = ["Partner","Activity Type","Location","District","Start","End","Status","Beneficiaries","Description"]
                rows_xl = [
                    [a["partner_name"], a["activity_type"], a["location"],
                     a.get("district",""), a.get("start_date",""), a.get("end_date",""),
                     a["status"], a.get("beneficiaries_reached",0), a.get("description","")]
                    for a in activities_raw
                ]
                st.download_button(
                    "🗺️ 4W Matrix (Excel)",
                    make_excel({"4W Matrix": (headers, rows_xl)}),
                    file_name=f"4W_Matrix_{date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

        # Full Situation Report Excel
        with col5:
            if sitreps:
                headers = ["Date","Day","Cases","Deaths","CFR%","HCW Affected","ETUs","Funding (USD)","Gap (USD)","Notes","Created By"]
                rows_xl = [
                    [s["report_date"], s["outbreak_day"], s["confirmed_cases"],
                     s["deaths"], s["cfr_percent"], s["healthcare_workers_affected"],
                     s["etus_operational"], s["total_funding_mobilized"],
                     s["funding_gap"], s.get("notes",""), s.get("created_by","")]
                    for s in sitreps
                ]
                st.download_button(
                    "📄 Situation Reports (Excel)",
                    make_excel({"SitReps": (headers, rows_xl)}),
                    file_name=f"SitRep_Summary_{date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

        # Full multi-sheet Excel report
        st.divider()
        if st.button("Generate Full Multi-Sheet Excel Report"):
            sheets = {}
            if partners:
                sheets["Partners"] = (
                    ["Name","Type","Country","Contact","Email","Status"],
                    [[p["name"],p["partner_type"],p["country"],
                      p.get("contact_person",""),p.get("contact_email",""),p["status"]]
                     for p in partners],
                )
            if resources:
                sheets["Resources"] = (
                    ["Partner","Type","Amount","Currency","Status"],
                    [[r["partner_name"],r["resource_type"],r["amount"],r["currency"],r["status"]]
                     for r in resources],
                )
            if activities_raw:
                sheets["Activities"] = (
                    ["Partner","Type","Location","Status","Beneficiaries"],
                    [[a["partner_name"],a["activity_type"],a["location"],a["status"],
                      a.get("beneficiaries_reached",0)] for a in activities_raw],
                )
            if sitreps:
                sheets["SitReps"] = (
                    ["Date","Day","Cases","Deaths","CFR%"],
                    [[s["report_date"],s["outbreak_day"],s["confirmed_cases"],
                      s["deaths"],s["cfr_percent"]] for s in sitreps],
                )
            if pillars:
                sheets["Funding Gap"] = (
                    ["Pillar","Required","Funded","Gap","Coverage%"],
                    [[p["pillar_name"],p["amount_required_usd"],p["amount_funded_usd"],
                      p["gap_usd"],p["coverage_pct"]] for p in pillars],
                )
            if sheets:
                st.download_button(
                    "Download Full Excel Report",
                    make_excel(sheets),
                    file_name=f"EVD_Response_Full_{date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

